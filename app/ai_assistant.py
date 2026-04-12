"""
Agent mode — Claude AI 助理
透過自然語言查詢 Ragic 資料、分析銷售/庫存，並可匯出 Excel。
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

# 確保 app/ 在 Python 搜尋路徑內，無論從哪個目錄執行
_APP_DIR = Path(__file__).resolve().parent
if str(_APP_DIR) not in sys.path:
    sys.path.insert(0, str(_APP_DIR))

import anthropic
import openpyxl
import questionary
from rich.console import Console
from rich.rule import Rule
from rich.text import Text

# 從同目錄的 ragic_upload 匯入
from ragic_upload import (
    ragic_get,
    load_customers,
    SALES_ORDER_SHEET,
    DELIVERY_ORDER_SHEET,
    INVENTORY_SHEET,
    CUSTOMER_SHEET,
)

console = Console()

# ── 工具定義 ──────────────────────────────────────────────────

TOOLS = [
    {
        "name": "query_sales_orders",
        "description": "查詢 Ragic 銷貨單列表。可依日期範圍和狀態篩選。回傳最近的訂單資料。",
        "input_schema": {
            "type": "object",
            "properties": {
                "date_from": {
                    "type": "string",
                    "description": "開始日期 YYYY/MM/DD，可省略"
                },
                "date_to": {
                    "type": "string",
                    "description": "結束日期 YYYY/MM/DD，可省略"
                },
                "status": {
                    "type": "string",
                    "description": "訂單狀態篩選，例如：未出貨、預接單。可省略表示全部"
                },
                "limit": {
                    "type": "integer",
                    "description": "最多回傳幾筆，預設 100"
                }
            },
        },
    },
    {
        "name": "query_order_items",
        "description": "查詢特定銷貨單的商品明細（子表）。",
        "input_schema": {
            "type": "object",
            "properties": {
                "order_id": {
                    "type": "string",
                    "description": "Ragic 銷貨單的 record ID（數字）"
                }
            },
            "required": ["order_id"],
        },
    },
    {
        "name": "query_inventory",
        "description": "查詢倉庫庫存資料。",
        "input_schema": {
            "type": "object",
            "properties": {
                "warehouse_code": {
                    "type": "string",
                    "description": "倉庫代碼，例如 TW01。可省略表示全部倉庫"
                }
            },
        },
    },
    {
        "name": "query_customers",
        "description": "查詢客戶列表。",
        "input_schema": {
            "type": "object",
            "properties": {
                "keyword": {
                    "type": "string",
                    "description": "搜尋關鍵字（客戶名稱或代碼）。可省略表示全部"
                }
            },
        },
    },
    {
        "name": "query_delivery_orders",
        "description": "查詢出貨單列表。",
        "input_schema": {
            "type": "object",
            "properties": {
                "status": {
                    "type": "string",
                    "description": "篩選狀態。可省略表示全部"
                }
            },
        },
    },
    {
        "name": "export_to_excel",
        "description": "將資料匯出為 Excel 檔案，儲存到指定路徑。",
        "input_schema": {
            "type": "object",
            "properties": {
                "headers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "欄位標題列表"
                },
                "rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "資料列，每列為一個陣列，順序對應 headers"
                },
                "filename": {
                    "type": "string",
                    "description": "檔案名稱（不含副檔名），例如：銷售分析_202604"
                },
                "folder_path": {
                    "type": "string",
                    "description": "儲存資料夾路徑。省略時預設為桌面"
                }
            },
            "required": ["headers", "rows", "filename"],
        },
    },
]

SYSTEM_PROMPT = """你是潮玩波普國際有限公司（Boptoys）的 ERP 助理，專門協助分析和查詢 Ragic 資料庫中的銷售、庫存、客戶資料。

請依照用戶使用的語言回答（繁體中文或英文）。

你的主要能力：
- 查詢銷貨單、出貨單、庫存、客戶資料
- 分析銷售趨勢、客戶下單頻率、SKU 銷量排名
- 協助採購決策（根據歷史銷量建議補貨數量）
- 將分析結果匯出為 Excel

注意事項：
- 查詢資料時優先使用工具取得最新資訊
- 分析時要有具體數字支撐
- 如果用戶要求建立訂單、拋轉等寫入操作，請引導他們使用主選單的對應功能
- 日期格式使用 YYYY/MM/DD
"""

# ── 工具執行 ──────────────────────────────────────────────────

def _query_sales_orders(date_from=None, date_to=None, status=None, limit=100):
    data = ragic_get(SALES_ORDER_SHEET, limit=min(limit, 2000))
    results = []
    for rid, rec in data.items():
        order_date = rec.get("訂單日期", rec.get("日期", ""))
        order_status = rec.get("訂單狀態", rec.get("狀態", ""))
        if date_from and order_date and order_date < date_from:
            continue
        if date_to and order_date and order_date > date_to:
            continue
        if status and status not in order_status:
            continue
        results.append({
            "id": rid,
            "訂單編號": rec.get("訂單編號", ""),
            "客戶": rec.get("客戶", rec.get("客戶名稱", "")),
            "日期": order_date,
            "狀態": order_status,
            "小計": rec.get("小計", rec.get("訂單小計", "")),
            "稅額": rec.get("稅額", ""),
            "總計": rec.get("總計", rec.get("訂單總計", "")),
            "備註": rec.get("備註", ""),
        })
    return json.dumps({"count": len(results), "orders": results[:limit]}, ensure_ascii=False)


def _query_order_items(order_id):
    data = ragic_get(f"{SALES_ORDER_SHEET}/{order_id}")
    return json.dumps(data, ensure_ascii=False)


def _query_inventory(warehouse_code=None):
    data = ragic_get(INVENTORY_SHEET)
    results = []
    for rid, rec in data.items():
        wh = rec.get("倉庫代碼", "")
        if warehouse_code and warehouse_code not in wh:
            continue
        results.append({
            "id": rid,
            "倉庫代碼": wh,
            "倉庫名稱": rec.get("倉庫名稱", ""),
            "商品編號": rec.get("商品編號", ""),
            "商品名稱": rec.get("商品名稱", ""),
            "庫存編號": rec.get("庫存編號", ""),
            "數量": rec.get("數量", ""),
            "規格": rec.get("規格", ""),
            "單位": rec.get("單位", ""),
        })
    return json.dumps({"count": len(results), "inventory": results}, ensure_ascii=False)


def _query_customers(keyword=None):
    customers = load_customers()
    if keyword:
        customers = [c for c in customers if keyword in c.get("name", "") or keyword in c.get("code", "")]
    return json.dumps({"count": len(customers), "customers": customers[:200]}, ensure_ascii=False)


def _query_delivery_orders(status=None):
    data = ragic_get(DELIVERY_ORDER_SHEET)
    results = []
    for rid, rec in data.items():
        order_status = rec.get("狀態", "")
        if status and status not in order_status:
            continue
        results.append({
            "id": rid,
            "出貨單編號": rec.get("出貨單編號", ""),
            "客戶": rec.get("客戶", rec.get("客戶名稱", "")),
            "日期": rec.get("日期", rec.get("出貨日期", "")),
            "狀態": order_status,
        })
    return json.dumps({"count": len(results), "delivery_orders": results}, ensure_ascii=False)


def _export_to_excel(headers, rows, filename, folder_path=None):
    if folder_path:
        folder = Path(folder_path).expanduser()
    else:
        folder = Path.home() / "Desktop"
    folder.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = folder / f"{filename}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active

    # 表頭
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="C5A059")

    # 資料列
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自動欄寬
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(out_path)
    return json.dumps({"status": "success", "path": str(out_path)}, ensure_ascii=False)


def _execute_tool(tool_name: str, tool_input: dict) -> str:
    try:
        if tool_name == "query_sales_orders":
            return _query_sales_orders(**tool_input)
        elif tool_name == "query_order_items":
            return _query_order_items(**tool_input)
        elif tool_name == "query_inventory":
            return _query_inventory(**tool_input)
        elif tool_name == "query_customers":
            return _query_customers(**tool_input)
        elif tool_name == "query_delivery_orders":
            return _query_delivery_orders(**tool_input)
        elif tool_name == "export_to_excel":
            return _export_to_excel(**tool_input)
        return json.dumps({"error": f"未知工具：{tool_name}"})
    except Exception as e:
        return json.dumps({"error": str(e)})


# ── 主對話迴圈 ────────────────────────────────────────────────

_ANTHROPIC_KEY_FILE = Path.home() / ".boptoys-anthropic_key"


def _get_anthropic_api_key() -> str:
    """取得 Anthropic API key：env → key 檔 → 提示輸入。"""
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if api_key:
        return api_key
    if _ANTHROPIC_KEY_FILE.exists():
        api_key = _ANTHROPIC_KEY_FILE.read_text(encoding="utf-8").strip()
        if api_key:
            os.environ["ANTHROPIC_API_KEY"] = api_key
            return api_key
    console.print("[#B0A898]尚未設定 Anthropic API Key[/#B0A898]")
    console.print("[dim]請至 console.anthropic.com → API Keys 取得[/dim]")
    api_key = questionary.password("請輸入 Anthropic API Key：").ask() or ""
    if not api_key:
        console.print("[red]未輸入 API Key，返回主選單[/red]")
        return ""
    _ANTHROPIC_KEY_FILE.write_text(api_key, encoding="utf-8")
    os.environ["ANTHROPIC_API_KEY"] = api_key
    console.print(f"[#5A9A4A]✓ API Key 已儲存，下次不需再輸入[/#5A9A4A]")
    return api_key


def run_agent_mode():
    api_key = _get_anthropic_api_key()
    if not api_key:
        return

    client = anthropic.Anthropic(api_key=api_key)
    messages = []

    console.print(Rule(style="#C5A059"))
    console.print(Text("  ◈ Agent Mode | 智慧決策助手", style="bold #C5A059"))
    console.print(Rule(style="#C5A059"))
    console.print()
    console.print("  [bold #C5A059]▸ 即時查詢 (Real-time Query)[/bold #C5A059]")
    console.print("    [dim]└ 「TW01 現在還有多少 BBB042？」[/dim]")
    console.print("    [dim]└ 「幫我查所有倉庫的庫存狀況」[/dim]")
    console.print()
    console.print("  [bold #C5A059]▸ 數據決策 (Intelligence & Analysis)[/bold #C5A059]")
    console.print("    [dim]└ 「1～3 月銷售前五名 SKU 是哪些？」[/dim]")
    console.print("    [dim]└ 「分析 XXX 客戶的訂單趨勢與缺貨風險」[/dim]")
    console.print()
    console.print("  [bold #C5A059]▸ 供應鏈優化 (Supply Chain Suggestion)[/bold #C5A059]")
    console.print("    [dim]└ 「BBB042 建議補貨量（基於過去 3 個月銷量）」[/dim]")
    console.print("    [dim]└ 「列出當前庫存不足的清單」[/dim]")
    console.print()
    console.print("  [bold #C5A059]▸ 自動化報表 (Report Automation)[/bold #C5A059]")
    console.print("    [dim]└ 「整理本月未出貨訂單並存至 Excel」[/dim]")
    console.print("    [dim]└ 「匯出庫存低於預警線的商品清單」[/dim]")
    console.print()
    console.print("  [dim]輸入「退出」/ 'exit' 離開  ·  輸入「重設 key」/ 'reset key' 更換 API Key[/dim]")
    console.print(Rule(style="#3a3a3a"))
    console.print()

    from prompt_toolkit import PromptSession
    from prompt_toolkit.history import InMemoryHistory
    _session = PromptSession(history=InMemoryHistory())

    while True:
        try:
            console.print(Rule(style="#3a3a3a"))
            user_input = _session.prompt("❯  ").strip()
            console.print(Rule(style="#3a3a3a"))
        except (KeyboardInterrupt, EOFError):
            break

        if not user_input:
            continue
        if user_input.lower() in ("退出", "exit", "quit", "bye", "q"):
            break
        if user_input.lower() in ("重設 key", "reset key", "重設key", "resetkey"):
            _ANTHROPIC_KEY_FILE.unlink(missing_ok=True)
            os.environ.pop("ANTHROPIC_API_KEY", None)
            console.print("[#B0A898]已清除 API Key[/#B0A898]")
            api_key = _get_anthropic_api_key()
            if not api_key:
                break
            client = anthropic.Anthropic(api_key=api_key)
            messages = []
            continue

        messages.append({"role": "user", "content": user_input})

        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=4096,
                system=SYSTEM_PROMPT,
                tools=TOOLS,
                messages=messages,
            )

            # 工具呼叫迴圈
            while response.stop_reason == "tool_use":
                tool_results = []
                for block in response.content:
                    if block.type == "tool_use":
                        console.print(f"  [dim]→ 查詢中：{block.name}...[/dim]")
                        result = _execute_tool(block.name, block.input)
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": result,
                        })

                messages.append({"role": "assistant", "content": response.content})
                messages.append({"role": "user", "content": tool_results})

                response = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=4096,
                    system=SYSTEM_PROMPT,
                    tools=TOOLS,
                    messages=messages,
                )

            # 取得最終文字回應
            reply = ""
            for block in response.content:
                if hasattr(block, "text"):
                    reply += block.text

            messages.append({"role": "assistant", "content": response.content})
            console.print()
            console.print(f"[bold #C5A059]Claude[/bold #C5A059]  {reply}")
            console.print()

        except anthropic.APIError as e:
            console.print(f"[red]API 錯誤：{e}[/red]")
        except Exception as e:
            console.print(f"[red]錯誤：{e}[/red]")

    console.print(f"[#B0A898]已離開 Agent mode[/#B0A898]")
