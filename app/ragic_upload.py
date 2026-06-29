#!/usr/bin/env python3
"""
Ragic 銷貨單自動化上傳腳本
用法：
  python ragic_upload.py client/LE/0324T221.xlsx (檔案名稱)
  python ragic_upload.py --dry-run client/LE/0324T221.xlsx
"""

import argparse
import hashlib
import json
import logging
import os
import re
import shutil
import sys
import time
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional

from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parent.parent / ".env")

_VERSION_FILE = Path(__file__).resolve().parent.parent / "VERSION"
APP_VERSION = _VERSION_FILE.read_text().strip() if _VERSION_FILE.exists() else "?"

import questionary
import requests

QSTYLE = questionary.Style([
    ("question",                        "bold #D4C9B0"),
    ("answer",                          "fg:#5A9A4A bold"),
    ("pointer",                         "fg:#FF7700 bold"),
    ("highlighted",                     "bg:#C5A059 fg:#1A1A1A bold"),
    ("text",                            "fg:#D4C9B0"),
    ("instruction",                     "fg:#666666"),
    ("checkbox",                        "fg:#C5A059"),
    ("checkbox-selected",               "fg:#FF7700 bold"),
    # autocomplete 下拉選單
    ("completion-menu.completion",          "bg:#2a2a2a fg:#D4C9B0"),
    ("completion-menu.completion.current",  "bg:#C5A059 fg:#1A1A1A bold"),
])

# 全域套用 QSTYLE，所有 questionary 呼叫自動帶入樣式
def _q_styled(fn):
    def wrapper(*args, **kwargs):
        kwargs.setdefault("style", QSTYLE)
        return fn(*args, **kwargs)
    return wrapper

questionary.select      = _q_styled(questionary.select)
questionary.checkbox    = _q_styled(questionary.checkbox)
questionary.confirm     = _q_styled(questionary.confirm)
questionary.text        = _q_styled(questionary.text)
questionary.password    = _q_styled(questionary.password)
questionary.autocomplete = _q_styled(questionary.autocomplete)


def _select_with_esc(message: str, choices: list):
    """questionary.select wrapper：Esc 鍵視為取消，回傳 None 以觸發退出。"""
    from prompt_toolkit.key_binding import KeyBindings, merge_key_bindings
    q = questionary.select(message, choices=choices)
    try:
        kb = KeyBindings()
        @kb.add("escape")
        def _(event):
            event.app.exit(result=None)
        q.application.key_bindings = merge_key_bindings([q.application.key_bindings, kb])
    except Exception:
        pass
    return q.ask()


def _pause():
    """操作完成後等待用戶按 Enter，避免結果訊息被主選單蓋掉。"""
    try:
        input("\n[按 Enter 返回主選單]")
    except (KeyboardInterrupt, EOFError):
        pass


from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.rule import Rule
from rich.text import Text

# ============================================================
# ★ CONFIG ★
# ============================================================

RAGIC_BASE    = os.getenv("RAGIC_BASE",    "https://ap12.ragic.com")
RAGIC_ACCOUNT = os.getenv("RAGIC_ACCOUNT", "toybebop")

PRODUCT_PRICE_SHEET  = os.getenv("PRODUCT_PRICE_SHEET",  "ragicsales-order-management/20006")  # 商品單價管理
CUSTOMER_SHEET       = os.getenv("CUSTOMER_SHEET",       "ragicsales-order-management/20004")  # 客戶
# 客戶表(20004)欄位 ID（CID）。Ragic 寫入 API 只認 CID 不認名稱；可用 ?api&naming=EID 取得。
CUSTOMER_FIELD_CIDS = {
    "name":      "3000479",  # 客戶名稱（必填）
    "short":     "3001873",  # 客戶簡稱
    "owner":     "3000480",  # 客戶負責人
    "contact":   "3001449",  # 主要聯絡窗口
    "mobile":    "3000909",  # 窗口手機
    "phone":     "3000483",  # 電話號碼
    "ship_addr": "3000903",  # 送貨地址
    "remark":    "3000913",  # 備註
    "code":      "3000666",  # 客戶編號（自動產生，唯讀，不可填）
}
SALES_ORDER_SHEET    = os.getenv("SALES_ORDER_SHEET",    "ragicsales-order-management/20001")  # 銷貨單
DELIVERY_ORDER_SHEET = os.getenv("DELIVERY_ORDER_SHEET", "ragicsales-order-management/20002")  # 出貨單
OUTBOUND_ORDER_SHEET = os.getenv("OUTBOUND_ORDER_SHEET", "ragicinventory/20009")               # 出庫單
INVENTORY_SHEET      = os.getenv("INVENTORY_SHEET",      "ragicinventory/20008")               # 倉庫庫存
STOCK_CHECK_WAREHOUSE = os.getenv("STOCK_CHECK_WAREHOUSE", "TW01")  # 開單前庫存把關的出貨倉（預設台灣總部）

ORDER_ITEMS_SUBTABLE_KEY    = os.getenv("ORDER_ITEMS_SUBTABLE_KEY",    "_subtable_3000842")  # 銷貨單訂購項目子表
OUTBOUND_ITEMS_SUBTABLE_KEY = os.getenv("OUTBOUND_ITEMS_SUBTABLE_KEY", "_subtable_3001132")  # 出庫單項目子表
OUTBOUND_DOC_NOTE_CID  = "3001121"  # 出庫單「單據備註」（表頭）→ 自動填客戶名稱
OUTBOUND_ROW_NOTE_CID  = "3001128"  # 出庫單子表「備註」（每列）→ 一律填國際條碼
INVENTORY_QTY_CID      = "3001107"  # 倉庫庫存 20008「數量」→ 自動拆盒時直接改

# 批次發樣：組合範本與客戶名單存本機 JSON（行政只在選單操作，不碰檔）
SAMPLE_COMBOS_FILE   = os.path.join(os.path.dirname(__file__), "sample_combos.json")
SAMPLE_CUSTLIST_FILE = os.path.join(os.path.dirname(__file__), "sample_customer_lists.json")
SAMPLE_ORDER_TYPES   = ["樣品申請", "公關品", "活動贈品"]

# 客戶尚未建檔時使用的預留客戶
UNREGISTERED_CUSTOMER = {"code": "C-00000", "name": "000尚未建檔", "address": ""}

# 上傳記錄檔（防重複）
_UPLOAD_LOG = Path(__file__).resolve().parent.parent / "upload_log.json"

# 操作日誌資料夾
_LOGS_DIR = Path(__file__).resolve().parent.parent / "logs"


def _load_upload_log() -> dict:
    if _UPLOAD_LOG.exists():
        try:
            return json.loads(_UPLOAD_LOG.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _save_upload_log(log: dict):
    _UPLOAD_LOG.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")


def _setup_logging():
    _LOGS_DIR.mkdir(exist_ok=True)
    log_file = _LOGS_DIR / f"{date.today()}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8")],
    )

_CID_LABELS = {
    "3000812": "訂單單別",    "3000813": "訂單日期",      "3000814": "訂單狀態",
    "3000815": "客戶編號",    "3000836": "課稅別",        "3000838": "稅率",
    "3001498": "訂單運費",    "3001684": "國貿條規",      "3000835": "小計",
    "3000837": "稅額",        "3000839": "總金額(含稅)",  "3000840": "備註",
    "1000074": "內部備注",    "3000845": "建檔日期時間",  "3000847": "最後修改日期時間",
    "3000830": "商品販售代號","3000832": "單價",          "3000833": "數量",
    "3000834": "金額",
}


def _humanize_payload(payload: dict) -> dict:
    result = {}
    for k, v in payload.items():
        if k == ORDER_ITEMS_SUBTABLE_KEY:
            rows = {rk: {_CID_LABELS.get(ck, ck): cv for ck, cv in rv.items()}
                    for rk, rv in v.items()}
            result["訂購項目"] = rows
        else:
            result[_CID_LABELS.get(k, k)] = v
    return result


# ============================================================

console = Console()

_KEY_FILE = Path.home() / ".boptoys-ai_key"

# ── Ragic API ────────────────────────────────────────────────

def _get_api_key(force_prompt: bool = False) -> str:
    """取得 Ragic API Key；force_prompt=True 時忽略既有金鑰，直接要求重新輸入。"""
    api_key = "" if force_prompt else os.environ.get("RAGIC_API_KEY", "")
    if not api_key and not force_prompt and _KEY_FILE.exists():
        api_key = _KEY_FILE.read_text().strip()
        os.environ["RAGIC_API_KEY"] = api_key
    if not api_key:
        if not force_prompt:
            console.print("[#FF7700]尚未設定 RAGIC_API_KEY[/#FF7700]")
        api_key = questionary.password("請輸入 Ragic API Key：").ask() or ""
        if not api_key:
            console.print("[red]未輸入 API Key，結束[/red]")
            sys.exit(1)
        _KEY_FILE.write_text(api_key, encoding="utf-8")
        os.environ["RAGIC_API_KEY"] = api_key
        console.print(f"[#5A9A4A]✓ API Key 已儲存至 {_KEY_FILE}，下次不需再輸入[/#5A9A4A]")
    return api_key


def _invalidate_api_key() -> None:
    """清除目前（失效的）API Key，下次會要求重新輸入。"""
    os.environ.pop("RAGIC_API_KEY", None)
    try:
        _KEY_FILE.unlink()
    except FileNotFoundError:
        pass


def _auth_header() -> dict:
    """Ragic API key 已是 Base64 格式，直接作為 Basic auth token。"""
    return {"Authorization": f"Basic {_get_api_key()}"}


def _is_auth_error(data) -> bool:
    """判斷 Ragic 回應是否為「金鑰失效／無存取權」錯誤（HTTP 200 但 body 是 ERROR）。"""
    return (
        isinstance(data, dict)
        and data.get("status") == "ERROR"
        and (data.get("code") == 106 or "guest account" in str(data.get("msg", "")))
    )


def _ragic_request(method: str, url: str, **kwargs) -> requests.Response:
    """帶自動重試的 HTTP 請求（最多 3 次，指數退避 1s/2s/4s）。"""
    retryable_errors = (requests.exceptions.ConnectionError, requests.exceptions.Timeout)
    last_exc = None
    for attempt in range(3):
        try:
            r = requests.request(method, url, **kwargs)
            if r.status_code >= 500 and attempt < 2:
                wait = 2 ** attempt
                console.print(f"[#FF7700]⚠ 伺服器錯誤（{r.status_code}），{wait} 秒後重試...[/#FF7700]")
                logging.warning("HTTP %s on %s, retrying in %ss (attempt %d)", r.status_code, url, wait, attempt + 1)
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r
        except retryable_errors as e:
            last_exc = e
            if attempt < 2:
                wait = 2 ** attempt
                console.print(f"[#FF7700]⚠ 網路錯誤，{wait} 秒後重試...[/#FF7700]")
                logging.warning("Network error on %s: %s, retrying in %ss (attempt %d)", url, e, wait, attempt + 1)
                time.sleep(wait)
    raise last_exc


def _ragic_json(method: str, url: str, *, timeout: int = 30, **kwargs) -> dict:
    """送出 Ragic API 請求並回傳 JSON；偵測到金鑰失效時清除舊金鑰、要求重填後重試一次。"""
    for attempt in range(2):
        r = _ragic_request(method, url, headers=_auth_header(), timeout=timeout, **kwargs)
        data = r.json()
        if _is_auth_error(data):
            if attempt == 0:
                console.print(
                    f"[#FF7700]⚠ Ragic API Key 失效或無此表存取權（{data.get('msg', '')}）[/#FF7700]"
                )
                _invalidate_api_key()
                _get_api_key(force_prompt=True)
                continue
            raise RuntimeError(f"Ragic API 存取失敗：{data.get('msg', data)}")
        return data
    raise RuntimeError("Ragic API 存取失敗")


def ragic_get(sheet_path: str, limit: int = 2000) -> dict:
    url = f"{RAGIC_BASE}/{RAGIC_ACCOUNT}/{sheet_path}?api&limit={limit}"
    data = _ragic_json("GET", url)
    return {k: v for k, v in data.items() if not k.startswith("_") and k != "info"}


def ragic_post(sheet_path: str, payload: dict) -> dict:
    url = f"{RAGIC_BASE}/{RAGIC_ACCOUNT}/{sheet_path}?api&doLinkLoad=true"
    return _ragic_json("POST", url, json=payload)


def ragic_patch(sheet_path: str, record_id: str, payload: dict) -> dict:
    url = f"{RAGIC_BASE}/{RAGIC_ACCOUNT}/{sheet_path}/{record_id}?api&doLinkLoad=true"
    return _ragic_json("PATCH", url, json=payload)


def ragic_get_action_button_id(sheet_path: str, button_name: str) -> Optional[int]:
    """從 Ragic metadata 取得指定名稱的 action button ID（massOperation 類別）。"""
    url = f"{RAGIC_BASE}/{RAGIC_ACCOUNT}/{sheet_path}/metadata/actionButton?api&category=massOperation"
    data = _ragic_json("GET", url)
    for btn in data.get("actionButtons", []):
        if btn.get("name") == button_name:
            return btn["id"]
    return None


def ragic_trigger_button(sheet_path: str, record_id: str, button_id) -> dict:
    """對單筆記錄觸發 Ragic action button。"""
    url = f"{RAGIC_BASE}/{RAGIC_ACCOUNT}/{sheet_path}/{record_id}?api&bId={button_id}"
    return _ragic_json("POST", url, timeout=60)


def _friendly_error(msg: str) -> str:
    """將常見 Ragic 錯誤訊息轉換為友善中文說明。"""
    m = str(msg).lower()
    if "already exist" in m or "duplicate" in m or "重複" in m:
        return "此記錄已存在（Ragic 擋掉重複建立），請至 Ragic 確認是否已建單"
    if "access right" in m or "permission" in m or "403" in m:
        return "權限不足，請確認 API Key 是否有此表單的存取權限"
    if "invalid" in m and "key" in m:
        return "API Key 無效，請重新設定（執行時選「重設 key」）"
    if "timeout" in m or "timed out" in m:
        return "連線逾時，請確認網路狀況後重試"
    if "connect" in m or "connection" in m:
        return "無法連線至 Ragic，請確認網路是否正常"
    if "not found" in m or "404" in m:
        return "找不到對應記錄，可能已被刪除或編號有誤"
    return msg


# ── 快取載入 ─────────────────────────────────────────────────

def load_price_index() -> Dict[str, list]:
    """載入商品單價管理，建立 {條碼: [商品...]} 索引。"""
    with console.status("[#B0A898]載入商品單價資料...[/#B0A898]", spinner="dots"):
        records = ragic_get(PRODUCT_PRICE_SHEET)
    index: Dict[str, list] = {}
    for rec in records.values():
        barcode = str(rec.get("國際條碼", "")).strip()
        if len(barcode) < 12:
            continue
        entry = {
            "product_code": str(rec.get("商品單價代號", "")),
            "product_name": str(rec.get("商品名稱", "")),
            "spec":         rec.get("規格", 1),
            "unit":         str(rec.get("單位", "")),
            "price":        float(rec.get("價格", 0) or 0),
        }
        index.setdefault(barcode, []).append(entry)
    console.print(f"[#5A9A4A]✓ 載入 {len(index)} 種商品[/#5A9A4A]")
    return index


_stock_cache: Dict[str, Dict[str, float]] = {}


def _to_int(x, default: int = 0) -> int:
    """容錯轉整數：吃全形數字、小數、空值。"""
    s = str(x).strip().translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return default


def extract_po(text: str) -> str:
    """從銷貨單備註抓 PO 號（如「【系統建單】 PO#53961 …」→ 53961）。找不到回空字串。
    必須有 PO# 井號才認（避免誤抓 polypeng 這類字串）；容忍井號後空格。"""
    m = re.search(r"PO\s*[#＃]\s*([A-Za-z0-9\-]+)", str(text or ""))
    return m.group(1) if m else ""


def compute_break_plan(line_needs: list, inventory: dict, warehouse_code: str) -> list:
    """自動拆盒計畫（唯讀，不寫入）。

    line_needs = [(來源標籤, 商品編號, 需求數量), ...]，**逐出貨單明細分開**分類，
    不同出貨單/客戶的同商品不會被合併（避免 TRU 18 + 蝦皮 1 = 19 被誤判成非整中盒）。
    inventory = 20008 全表 {rid: rec}。規格 = 每盒入數（單盒1、中盒N、箱M）；
    同家族＝去掉商品編號末碼（單盒1/中盒2/箱3）。

    規則：
    - 單盒線(規格=1)、需求整中盒 → 拆 需求÷入數 個中盒，現有散單盒不算（保留）。
    - 單盒線、非整中盒 → 零售/填錯：用現有散單盒，不夠才提醒拆實體（不自動）。
    - 中盒以上線 → 直接扣、不進清單。
    - 多單同時拆同一中盒 → 累計扣，逐單檢查中盒是否還夠。
    回傳每筆含 label（哪一單）、status：ok／parent_short／no_parent／no_stock／manual。
    """
    detail = {}
    for rid, rec in inventory.items():
        if str(rec.get("倉庫代碼", "")).strip() != warehouse_code:
            continue
        prod = str(rec.get("商品編號", "")).strip()
        if not prod:
            continue
        detail[prod] = {
            "spec": _to_int(rec.get("規格", 1), 1) or 1,
            "qty":  _to_int(rec.get("數量", 0), 0),
            "rid":  rid,
            "code": str(rec.get("庫存編號", "")).strip(),
            "name": str(rec.get("商品名稱", "")).strip(),
        }

    consumed = {}  # 中盒商品編號 → 本批已規劃拆出的盒數（跨單累計，檢查中盒夠不夠）
    plan = []
    for label, prod, q in line_needs:
        q = int(q)
        d = detail.get(prod)
        if not d:
            plan.append({"label": label, "prod": prod, "name": "", "need": q,
                         "have": 0, "status": "no_stock"})
            continue
        if d["spec"] != 1:
            continue  # 中盒（含以上）線：客戶用中盒下單，直接扣、不進拆盒清單
        fam = prod[:-1]
        parents = sorted(
            [(p, detail[p]) for p in detail
             if p != prod and p[:-1] == fam and detail[p]["spec"] > d["spec"]],
            key=lambda kv: kv[1]["spec"])
        if not parents:
            if d["qty"] >= q:
                continue
            plan.append({"label": label, "prod": prod, "name": d["name"],
                         "need": q, "have": d["qty"], "status": "no_parent"})
            continue
        pcode, pd = parents[0]  # 取最接近的上一級中盒
        ratio = pd["spec"]      # 單盒規格=1，故入數=中盒規格
        exact = (q % ratio == 0)
        entry = {
            "label": label, "prod": prod, "name": d["name"], "need": q, "have": d["qty"],
            "parent": pcode, "parent_name": pd["name"], "ratio": ratio,
            "unit_rid": d["rid"], "unit_qty": d["qty"],
            "parent_rid": pd["rid"], "parent_qty": pd["qty"], "exact": exact,
        }
        if exact:
            boxes = q // ratio
            avail = pd["qty"] - consumed.get(pcode, 0)  # 扣掉本批前面幾單已規劃拆的
            entry["boxes"] = boxes
            entry["gain"] = boxes * ratio
            if avail >= boxes:
                entry["status"] = "ok"
                consumed[pcode] = consumed.get(pcode, 0) + boxes
            else:
                entry["status"] = "parent_short"
                entry["avail"] = avail
        else:
            shortage = max(0, q - d["qty"])
            entry["boxes"] = -(-shortage // ratio) if shortage else 0
            entry["gain"] = entry["boxes"] * ratio
            entry["status"] = "manual"
        plan.append(entry)
    return plan


def build_code_to_barcode(price_index: Dict[str, list]) -> Dict[str, list]:
    """商品編號（如 PWF011）→ 國際條碼清單。一碼多條碼時保留多筆，呼叫端再決定。"""
    out: Dict[str, list] = {}
    for barcode, entries in price_index.items():
        for e in entries:
            base = _strip_code_suffix(str(e.get("product_code", "")))
            if not base:
                continue
            lst = out.setdefault(base, [])
            if barcode not in lst:
                lst.append(barcode)
    return out


def _strip_code_suffix(code: str) -> str:
    """商品單價代號(如 BMC012-1) → 庫存商品編號(BMC012)：剝掉結尾 -數字。"""
    return re.sub(r'-\d+$', '', str(code).strip())


def load_stock(warehouse: str = STOCK_CHECK_WAREHOUSE) -> Dict[str, float]:
    """載入指定倉庫的庫存 {商品編號: 數量}，模組級快取（一次 API）。
    庫存依規格分開記（中盒/單盒/整箱各一筆），故數量單位與該規格代號一致。"""
    if warehouse in _stock_cache:
        return _stock_cache[warehouse]
    with console.status(f"[#B0A898]載入 {warehouse} 庫存...[/#B0A898]", spinner="dots"):
        records = ragic_get(INVENTORY_SHEET)
    stock: Dict[str, float] = {}
    for rec in records.values():
        if str(rec.get("倉庫代碼", "")).strip() != warehouse:
            continue
        code = str(rec.get("商品編號", "")).strip()
        if not code:
            continue
        try:
            stock[code] = stock.get(code, 0.0) + float(rec.get("數量") or 0)
        except (ValueError, TypeError):
            pass
    _stock_cache[warehouse] = stock
    return stock


def check_stock(resolved: list, warehouse: str = STOCK_CHECK_WAREHOUSE) -> list:
    """比對每項的下單量 vs 出貨倉庫存，顯示對照表並標紅不足項。
    回傳不足項的說明字串清單（不阻擋，只提醒）。庫存無此代號者視為未追蹤、不判定。"""
    try:
        stock = load_stock(warehouse)
    except Exception as e:
        console.print(f"[#FF7700]⚠ 無法載入庫存，略過庫存把關：{e}[/#FF7700]")
        return []

    table = Table(show_header=True, header_style="bold #C5A059", box=None)
    table.add_column("商品名稱", min_width=20)
    table.add_column("規格", width=5, justify="right")
    table.add_column("下單", width=6, justify="right")
    table.add_column(f"{warehouse}庫存", width=8, justify="right")
    table.add_column("判定", width=10)

    shortages = []
    for it in resolved:
        base = _strip_code_suffix(it["product_code"])
        avail = stock.get(base)            # None = 庫存表未追蹤此代號
        ordered = it["quantity"]
        if avail is None:
            verdict, avail_txt = "[dim]無庫存資料[/dim]", "—"
        elif ordered > avail:
            short = ordered - avail
            verdict, avail_txt = f"[bold #D14040]✗ 缺 {short:g}[/bold #D14040]", f"{avail:g}"
            shortages.append(f"{it['product_name']} 下單{ordered:g} > {warehouse}庫存{avail:g}（缺{short:g}）")
        else:
            verdict, avail_txt = "[#5A9A4A]✓ 足[/#5A9A4A]", f"{avail:g}"
        name = it["product_name"][:22] + ("…" if len(it["product_name"]) > 22 else "")
        style = "#D14040" if (avail is not None and ordered > avail) else None
        table.add_row(name, str(it["spec"]), f"{ordered:g}", avail_txt, verdict, style=style)

    console.print(f"\n[bold]庫存把關（出貨倉 {warehouse}）[/bold]")
    console.print(table)
    if shortages:
        console.print(f"[bold #D14040]⚠ {len(shortages)} 項庫存不足，請確認是否仍要開單[/bold #D14040]")
    return shortages


def load_customers() -> list:
    """載入客戶資料表。"""
    with console.status("[#B0A898]載入客戶資料...[/#B0A898]", spinner="dots"):
        records = ragic_get(CUSTOMER_SHEET)
    customers = []
    for rec in records.values():
        customers.append({
            "code":    str(rec.get("客戶編號", "")),
            "name":    str(rec.get("客戶名稱", "")),
            "address": str(rec.get("送貨完整地址", "")),
        })
    console.print(f"[#5A9A4A]✓ 載入 {len(customers)} 筆客戶[/#5A9A4A]")
    return customers


# ── 客戶比對 ─────────────────────────────────────────────────

def create_customer_interactive(store_code: str, client_code: str, customers: list,
                                dry_run: bool = False) -> Optional[dict]:
    """互動式新建客戶並寫入 Ragic 客戶表(20004)。成功則加入 customers 快取並回傳。
    客戶編號由 Ragic 自動產生（不填）。取消/失敗回 None。"""
    # 客戶命名沒有單一規則（公司全名／個人名／通路前綴皆有）。
    # 從現有同通路客戶學前綴並顯示範例，讓使用者照既有慣例命名，而非硬套 client_code。
    def _channel_key(nm: str) -> str:
        return re.sub(r'[^A-Za-z]', '', str(nm).split("-")[0]).upper()
    code_key = re.sub(r'[^A-Za-z]', '', client_code or "").upper()
    channel_names = [c["name"] for c in customers
                     if code_key and _channel_key(c["name"]).startswith(code_key)]
    prefix = ""
    if channel_names and "-" in channel_names[0]:
        prefix = channel_names[0][:channel_names[0].index("-") + 1]   # 例：TRU- / L.E.-
    suggested = f"{prefix}{store_code}" if prefix else (
        f"{client_code}-{store_code}" if client_code else str(store_code))
    if channel_names:
        console.print(f"[dim]同通路現有命名參考：{'、'.join(channel_names[:5])}[/dim]")
    name = (questionary.text("客戶名稱（必填，可改）", default=suggested).ask() or "").strip()
    if not name:
        console.print("[#FF7700]未輸入客戶名稱，取消建立[/#FF7700]")
        return None
    short   = (questionary.text("客戶簡稱（門市名，可留空）", default="").ask() or "").strip()
    contact = (questionary.text("主要聯絡窗口（可留空）", default="").ask() or "").strip()
    phone   = (questionary.text("聯絡電話/手機（可留空）", default="").ask() or "").strip()
    addr    = (questionary.text("送貨地址（可留空）", default="").ask() or "").strip()
    owner   = (questionary.text("客戶負責人", default="Woody").ask() or "Woody").strip()

    console.print(Panel(
        f"客戶名稱：{name}\n客戶簡稱：{short or '—'}\n聯絡窗口：{contact or '—'}\n"
        f"電話/手機：{phone or '—'}\n送貨地址：{addr or '—'}\n負責人：{owner}",
        title="新客戶（客戶編號由 Ragic 自動產生）", border_style="#C5A059"))
    if not questionary.confirm("確認建立此客戶？", default=True).ask():
        console.print("[#FF7700]已取消建立[/#FF7700]")
        return None

    C = CUSTOMER_FIELD_CIDS
    payload = {C["name"]: name, C["owner"]: owner}
    if short:   payload[C["short"]]     = short
    if contact: payload[C["contact"]]   = contact
    if phone:   payload[C["mobile"]]    = phone; payload[C["phone"]] = phone
    if addr:    payload[C["ship_addr"]] = addr

    if dry_run:
        console.print("[#FF7700]★ DRY-RUN：不實際建立客戶[/#FF7700]")
        cust = {"code": "C-DRYRUN", "name": name, "address": addr}
    else:
        try:
            result = ragic_post(CUSTOMER_SHEET, payload)
        except Exception as e:
            console.print(f"[red]✗ 建立客戶失敗：{e}[/red]")
            return None
        if result.get("status") != "SUCCESS" and not result.get("ragicId"):
            console.print(f"[red]✗ 建立客戶失敗：{result.get('msg', result)}[/red]")
            return None
        new_code = str(result.get("data", {}).get(C["code"], "")).strip()
        console.print(f"[#5A9A4A]✓ 客戶建立成功！編號 {new_code}（Ragic ID {result.get('ragicId', '')}）[/#5A9A4A]")
        cust = {"code": new_code, "name": name, "address": addr}
    customers.append(cust)   # 加入快取，後續比對找得到
    return cust


def find_customer(customers: list, store_code: str, client_code: str = "",
                  dry_run: bool = False) -> Optional[dict]:
    # 若有 client_code（如 TRU），優先在該客群中搜尋
    if client_code:
        narrowed = [c for c in customers if store_code in c["name"] and client_code in c["name"]]
        if narrowed:
            matches = narrowed
        else:
            matches = [c for c in customers if store_code in c["name"]]
    else:
        matches = [c for c in customers if store_code in c["name"]]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        choices = [f"{c['code']}  {c['name']}" for c in matches]
        sel = questionary.select(f"找到多個含「{store_code}」的客戶，請選擇：", choices=choices).ask()
        return matches[choices.index(sel)]
    # 找不到 → 提供：建立新客戶 / 搜尋現有 / 暫用尚未建檔
    console.print(f"[#FF7700]⚠ 找不到含「{store_code}」的客戶[/#FF7700]")
    BUILD, SEARCH, PLACEHOLDER = "➕ 建立新客戶", "🔍 搜尋現有客戶", "暫用「C-00000 尚未建檔」"
    action = questionary.select("請選擇：", choices=[BUILD, SEARCH, PLACEHOLDER]).ask()
    if action == BUILD:
        new = create_customer_interactive(store_code, client_code, customers, dry_run)
        return new or UNREGISTERED_CUSTOMER
    if action == PLACEHOLDER or not action:
        return UNREGISTERED_CUSTOMER
    all_choices = [f"{c['code']}  {c['name']}" for c in customers]
    sel = questionary.autocomplete(
        "搜尋客戶（輸入代碼或名稱片段）：",
        choices=all_choices,
        validate=lambda v: v in all_choices or "請從清單中選擇",
        style=QSTYLE,
    ).ask()
    return customers[all_choices.index(sel)]


# ── 商品比對 ─────────────────────────────────────────────────

def _mid_pack_size(matches) -> Optional[int]:
    """從同條碼的多規格中找出「中盒/端盒」的每盒入數（pcs）。找不到回 None。"""
    for m in matches:
        unit = str(m.get("unit", ""))
        if "中盒" in unit or "端盒" in unit:
            try:
                n = int(float(m["spec"]))
                if n > 1:
                    return n
            except (ValueError, TypeError):
                pass
    return None


def resolve_items(order_items, price_index: dict, auto_unit_spec: bool = False) -> list:
    """
    auto_unit_spec=True：數量單位為「個/盒」時自動選 spec=1（LE 格式適用）

    另外：TRU 等以 PCS 下單的格式，若某 SKU 的 PCS 無法整除「中盒」入數，
    通常是客戶填錯，會在該項標記 box_note，最後寫進訂單內部備注提醒人工確認。
    """
    resolved = []
    for item in order_items:
        matches = price_index.get(item.barcode, [])
        if not matches:
            console.print(f"[#FF7700]⚠ 條碼 {item.barcode}（{item.le_name}）不在商品單價表，已跳過[/#FF7700]")
            continue

        # TRU 檔已帶單價，優先使用；否則從 Ragic 商品表取預設價
        override_price = item.unit_price if item.unit_price > 0 else None

        if len(matches) == 1:
            product = matches[0]
            final_qty = int(item.quantity)
        else:
            # 同條碼多規格：找出能整除數量的規格
            viable = []
            for m in matches:
                spec_qty = int(float(m["spec"])) if m["spec"] else 1
                if spec_qty <= 1 or item.quantity % spec_qty == 0:
                    n = int(item.quantity / spec_qty) if spec_qty > 1 else int(item.quantity)
                    viable.append((m, n))

            if not viable:
                viable = [(m, int(item.quantity)) for m in matches]

            # 按規格數值升冪排列：單盒(1) → 中盒 → 整箱
            viable.sort(key=lambda x: int(float(x[0]["spec"]) if x[0]["spec"] else 1))

            if len(viable) == 1:
                product, final_qty = viable[0]
            elif auto_unit_spec:
                # LE 格式：數量為個/盒單位，自動選 spec=1（最小單位）
                unit_options = [(m, n) for m, n in viable if int(float(m["spec"]) if m["spec"] else 1) == 1]
                if unit_options:
                    product, final_qty = unit_options[0]
                else:
                    product, final_qty = viable[0]
            else:
                choices = [
                    f"{m['unit']} × {n}"
                    + (f"  ({int(float(m['spec']))}pcs/盒)" if int(float(m['spec']) if m['spec'] else 1) > 1 else "")
                    + f"  ({m['product_code']} @ {m['price']:.2f} = {m['price']*n:,.2f})"
                    for m, n in viable
                ]
                name_hint = item.le_name or item.barcode
                sel = questionary.select(
                    f"{name_hint}（數量: {int(item.quantity)}）- 請選擇規格",
                    choices=choices,
                ).ask()
                product, final_qty = viable[choices.index(sel)]

        # 檢查 PCS 是否為整數中盒（僅對非 auto_unit_spec 的格式，如 TRU）。
        # 不整除多半是客戶填錯數量 → 標記，稍後寫入訂單內部備注提醒確認。
        box_note = ""
        if not auto_unit_spec:
            mid = _mid_pack_size(matches)
            if mid and item.quantity % mid != 0:
                box_note = f"{product['product_name']} {int(item.quantity)}pcs 非整中盒（{mid}個/中盒），請與客戶確認數量"
                console.print(f"[#FF7700]⚠ {box_note}[/#FF7700]")

        resolved.append({
            "product_code": product["product_code"],
            "product_name": product["product_name"],
            "spec":         product["spec"],
            "unit":         product["unit"],
            "unit_price":   override_price if override_price else product["price"],
            "quantity":     final_qty,
            "amount":       float((Decimal(str(override_price if override_price else product["price"])) * Decimal(str(final_qty))).quantize(Decimal("0.01"), ROUND_HALF_UP)),
            "box_note":     box_note,
        })
    return resolved


# ── 互動 UI ──────────────────────────────────────────────────

def show_items_table(customer: dict, store_code: str, po_number: str, resolved: list):
    console.print()
    console.rule(f"[bold]門市: {store_code}  PO: {po_number}  客戶: {customer['code']} {customer['name']}[/bold]")
    table = Table(show_header=True, header_style="bold #C5A059", box=None)
    table.add_column("#",       width=3)
    table.add_column("商品名稱", min_width=22)
    table.add_column("規格",    width=5,  justify="right")
    table.add_column("數量",    width=6,  justify="right")
    table.add_column("單價",    width=9,  justify="right")
    table.add_column("金額",    width=11, justify="right")
    subtotal = 0.0
    for i, it in enumerate(resolved, 1):
        table.add_row(str(i), it["product_name"], str(it["spec"]),
                      str(it["quantity"]), f"{it['unit_price']:,.2f}", f"{it['amount']:,.2f}")
        subtotal += it["amount"]
    console.print(table)
    console.print(f"[bold]小計: {subtotal:,.2f}[/bold]")
    console.print()


def ask_order_options(is_unregistered: bool = False) -> tuple:
    order_type = questionary.select(
        "請選擇訂單單別",
        choices=["一般訂單", "公關品", "樣品", "蝦皮", "官網"],
    ).ask()

    order_status = questionary.select(
        "請選擇訂單狀態",
        choices=["未出貨", "預接單", "已收款未出貨", "已出貨未收款", "尚未建檔"],
        default="尚未建檔" if is_unregistered else "未出貨",
    ).ask()

    tax_choice = questionary.select(
        "請選擇稅率",
        choices=["5%（含稅/外加）", "(5%)（內含/不計稅）"],
    ).ask()
    tax_rate = "5%" if "5%（" in tax_choice else "(5%)"

    shipping_str = questionary.text("運費（預設 0）", default="0").ask()
    shipping_fee = float(shipping_str or "0")

    commission = questionary.select(
        "業務分潤",
        choices=["8%", "2%", "（無）"],
        default="（無）",
    ).ask()
    commission = "" if commission == "（無）" else commission

    notes = questionary.text("備註（留空直接按 Enter）", default="").ask()
    internal_notes = questionary.text("內部備注（留空直接按 Enter）", default="").ask()
    return order_type, order_status, tax_rate, shipping_fee, notes or "", internal_notes or "", commission


def show_confirmation(customer: dict, resolved: list, order_type: str, order_status: str,
                      tax_rate: str, shipping_fee: float, notes: str, internal_notes: str,
                      commission: str = "") -> tuple:
    subtotal = sum(Decimal(str(it["amount"])) for it in resolved)
    tax_amount = (subtotal * Decimal("0.05")).quantize(Decimal("0.01"), ROUND_HALF_UP) if tax_rate == "5%" else Decimal("0")
    total = subtotal + tax_amount + Decimal(str(shipping_fee))

    console.print()
    console.rule("[bold red]最終確認[/bold red]")
    console.print(f"訂單單別: [#B0A898]{order_type}[/#B0A898]  狀態: [#B0A898]{order_status}[/#B0A898]  客戶: [#B0A898]{customer['code']}  {customer['name']}[/#B0A898]")
    console.print(f"課稅別: 營業稅              稅率: [#B0A898]{tax_rate}[/#B0A898]")
    console.print(f"小計: {subtotal:>12,.2f}     稅額: {tax_amount:,.2f}")
    console.print(f"運費: {shipping_fee:>12.2f}     總計: [bold]{total:,.2f}[/bold]")
    if commission:
        console.print(f"業務分潤: [#B0A898]{commission}[/#B0A898]")
    if notes:
        console.print(f"備註: {notes}")
    if internal_notes:
        console.print(f"內部備注: [dim]{internal_notes}[/dim]")
    console.rule()
    return subtotal, tax_amount, total


# ── Payload 組裝 ─────────────────────────────────────────────

def build_payload(customer: dict, resolved: list, order_type: str, order_status: str,
                  tax_rate: str, shipping_fee: float, notes: str, internal_notes: str,
                  commission: str = "") -> dict:
    now   = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    today = date.today().strftime("%Y/%m/%d")

    # 計算各項金額
    subtable = {}
    subtotal = Decimal("0")
    total_items = len(resolved)
    for i, it in enumerate(resolved):
        amount = (Decimal(str(it["unit_price"])) * Decimal(str(it["quantity"]))).quantize(Decimal("0.01"), ROUND_HALF_UP)
        subtotal += amount
        subtable[str(-(total_items - i))] = {
            "3000829": i + 1,                # 項次
            "3000830": it["product_code"],   # 商品販售代號
            "3000832": it["unit_price"],      # 單價
            "3000833": it["quantity"],        # 數量
            "3000834": float(amount),        # 金額（單價×數量）
        }

    tax_amount  = (subtotal * Decimal("0.05")).quantize(Decimal("0.01"), ROUND_HALF_UP) if tax_rate == "5%" else Decimal("0")
    total       = subtotal + tax_amount + Decimal(str(shipping_fee))

    # 非整中盒提醒：彙整各項 box_note，附加進內部備注（人工確認用，不印給客戶）。
    box_notes = [it.get("box_note") for it in resolved if it.get("box_note")]
    internal_parts = ["【程式建單】"]
    if internal_notes:
        internal_parts.append(internal_notes)
    if box_notes:
        internal_parts.append("⚠數量待確認：" + "；".join(box_notes))
    internal_full = " ".join(internal_parts)

    return {
        "3000812": order_type,               # 訂單單別
        "3000813": today,                    # 訂單日期
        "3000814": order_status,             # 訂單狀態
        "3000815": customer["code"],         # 客戶編號
        "3000836": "營業稅",                 # 課稅別
        "3000838": tax_rate,                 # 稅率
        "3001498": int(shipping_fee),        # 訂單運費
        "3001684": "DDP",                    # 國貿條規（預設）
        "3000835": float(subtotal),          # 小計
        "3000837": float(tax_amount),        # 稅額
        "3000839": float(total),             # 總金額(含稅)
        "3000840": notes,                    # 備註
        "1000065": commission,               # 業務分潤
        "1000074": internal_full,            # 內部備注（含非整中盒提醒）
        "3000845": now,                      # 建檔日期時間
        "3000847": now,                      # 最後修改日期時間
        ORDER_ITEMS_SUBTABLE_KEY: subtable,
    }


BASE_CLIENT_ORDER = Path(__file__).resolve().parent.parent / "client_order"
BASE_TEMPLATES    = Path(__file__).resolve().parent.parent / "templates"
BASE_OUTPUT       = Path(__file__).resolve().parent.parent / "exports"


def find_pending_files(base_dir: Path) -> list:
    files = []
    for client_dir in sorted(base_dir.iterdir()):
        if client_dir.is_dir():
            files.extend(sorted(client_dir.glob("*.xlsx")))
    return files


def process_file(excel_path: Path, args, price_index: dict, customers: list):
    from parsers import PARSERS
    client_code = excel_path.parent.name.upper()
    if client_code not in PARSERS:
        console.print(f"[red]不支援的客戶代碼：{client_code}（支援：{', '.join(PARSERS)}）[/red]")
        return 0, 0

    console.print(f"\n[#B0A898]解析 {excel_path.name}（{client_code} 格式）...[/#B0A898]")
    try:
        orders = PARSERS[client_code](str(excel_path)).parse()
    except Exception as e:
        console.print(f"[red]無法讀取 Excel 檔案：{e}[/red]")
        return 0, 0
    if not orders:
        console.print("[red]無法解析任何訂單，請確認檔案格式[/red]")
        return 0, 0
    console.print(f"[#5A9A4A]✓ 偵測到 {len(orders)} 張訂單[/#5A9A4A]")

    upload_log = _load_upload_log()
    file_hash = hashlib.md5(excel_path.read_bytes()).hexdigest()
    success_count = 0
    for i, order in enumerate(orders, 1):
        console.print(f"\n{'═'*58}")
        console.print(f"[bold]訂單 {i}/{len(orders)}  門市: {order.store_code}  PO: {order.po_number}[/bold]")

        # 防重複：以 PO 為單位判斷是否已上傳過
        log_key = f"{client_code}_{order.store_code}_{order.po_number}"
        if log_key in upload_log and not args.dry_run:
            rec = upload_log[log_key]
            console.print(f"[#FF7700]⚠ 此訂單已於 {rec['uploaded_at']} 上傳（Ragic ID: {rec['ragic_id']}）[/#FF7700]")
            logging.info("重複跳過 log_key=%s ragic_id=%s", log_key, rec['ragic_id'])
            skip = questionary.confirm("是否跳過（建議跳過以避免重複）？", default=True).ask()
            if skip:
                console.print("[#FF7700]已跳過[/#FF7700]")
                continue

        customer = find_customer(customers, order.store_code, client_code, dry_run=args.dry_run)
        if not customer:
            console.print("[red]無法確認客戶，跳過[/red]")
            continue

        resolved = resolve_items(order.items, price_index, auto_unit_spec=(client_code == "LE"))
        if not resolved:
            console.print("[red]無有效商品，跳過[/red]")
            continue

        show_items_table(customer, order.store_code, order.po_number, resolved)

        # 開單前庫存把關：比對出貨倉中盒庫存，不足標紅提醒（不阻擋）
        check_stock(resolved)

        is_unregistered = customer["code"] == UNREGISTERED_CUSTOMER["code"]
        order_type, order_status, tax_rate, shipping_fee, notes, internal_notes, commission = ask_order_options(is_unregistered)

        show_confirmation(customer, resolved, order_type, order_status, tax_rate, shipping_fee, notes, internal_notes, commission)

        confirmed = questionary.confirm("確認送出此訂單？", default=True).ask()
        if not confirmed:
            action = questionary.select(
                "請選擇：",
                choices=["跳過此單，繼續下一張", "放棄整個檔案，回到選單（不移至 done）"],
            ).ask()
            if "放棄" in action:
                if success_count > 0:
                    console.print(f"[#FF7700]已放棄。前 {success_count} 張已送出至 Ragic，請自行確認是否需要刪除。[/#FF7700]")
                else:
                    console.print("[#FF7700]已放棄，未送出任何訂單。[/#FF7700]")
                return success_count, len(orders), True
            console.print("[#FF7700]已跳過[/#FF7700]")
            continue

        payload = build_payload(customer, resolved, order_type, order_status, tax_rate, shipping_fee, notes, internal_notes, commission)
        console.print(json.dumps(_humanize_payload(payload), ensure_ascii=False, indent=2))

        if args.dry_run:
            console.print("[#FF7700]★ DRY-RUN，未實際送出[/#FF7700]")
            success_count += 1
        else:
            try:
                result = ragic_post(SALES_ORDER_SHEET, payload)
                if result.get("status") == "SUCCESS" or result.get("ragicId"):
                    ragic_id = result.get("ragicId", "")
                    console.print(f"[#5A9A4A]✓ 訂單建立成功！Ragic ID: {ragic_id}[/#5A9A4A]")
                    logging.info("銷貨單建立成功 ragic_id=%s file=%s log_key=%s", ragic_id, excel_path.name, log_key)
                    success_count += 1
                    upload_log[log_key] = {
                        "ragic_id":    str(ragic_id),
                        "uploaded_at": datetime.now().strftime("%Y/%m/%d %H:%M"),
                        "file":        excel_path.name,
                        "file_hash":   file_hash,
                    }
                    _save_upload_log(upload_log)
                else:
                    msg = result.get("msg", "") or str(result)
                    console.print(f"[red]✗ 建單失敗：{_friendly_error(msg)}[/red]")
            except Exception as e:
                console.print(f"[red]✗ 送出失敗：{_friendly_error(str(e))}[/red]")

    if success_count > 0 and not args.dry_run:
        done_dir = excel_path.parent / "done"
        done_dir.mkdir(exist_ok=True)
        dest = done_dir / excel_path.name
        shutil.move(str(excel_path), str(dest))
        console.print(f"[#5A9A4A]✓ 已移至 {dest.parent.name}/done/{dest.name}[/#5A9A4A]")
        logging.info("檔案移至 done: %s", dest)

    return success_count, len(orders), False


# ── 主選單流程 ───────────────────────────────────────────────

def run_new_sales_order(args, price_index: dict, customers: list):
    """新建銷售單（原 main while 迴圈，沒有 xlsx 時回主選單）。"""
    total_success = total_orders = 0
    while True:
        all_files = find_pending_files(BASE_CLIENT_ORDER)
        if not all_files:
            console.print("[#FF7700]沒有待處理的 Excel 檔案了，返回主選單[/#FF7700]")
            break

        labels = [f"{f.parent.name}/{f.name}" for f in all_files]
        selected = questionary.checkbox(
            "請選擇要處理的採購單（空白鍵勾選，Enter 確認；不選直接 Enter 返回）：",
            choices=[questionary.Choice(label, checked=False) for label in labels],
        ).ask()
        if not selected:
            console.print("[#FF7700]返回主選單[/#FF7700]")
            break

        excel_path = all_files[labels.index(selected[0])]

        console.print(f"[#B0A898]── 即將處理：{excel_path.parent.name}/{excel_path.name} ──[/#B0A898]")
        ok = questionary.confirm("確認執行？", default=True).ask()
        if not ok:
            continue

        s, o, _ = process_file(excel_path, args, price_index, customers)
        total_success += s
        total_orders  += o
        console.print(f"\n[bold cyan]{'─'*58}[/bold cyan]")

    if total_orders > 0:
        console.print(f"[bold #5A9A4A]本次共處理 {total_success}/{total_orders} 張訂單[/bold #5A9A4A]")


def run_create_delivery_order(args):
    """銷貨單批量拋轉建立出貨單（訂單狀態：未出貨 / 預接單 / 已收款未出貨）。"""
    TARGET_STATUSES = {"未出貨", "預接單", "已收款未出貨"}

    with console.status("[#B0A898]載入銷貨單資料...[/#B0A898]", spinner="dots"):
        records = ragic_get(SALES_ORDER_SHEET)

    candidates = []
    for rid, rec in records.items():
        status = str(rec.get("訂單狀態", ""))
        if status in TARGET_STATUSES:
            candidates.append({
                "id":    rid,
                "label": f"{rec.get('訂單編號','?')}  {rec.get('客戶名稱','?')}  {rec.get('訂單日期','?')}  [{status}]",
            })

    if not candidates:
        console.print("[#FF7700]沒有待拋轉的銷貨單（未出貨 / 預接單 / 已收款未出貨）[/#FF7700]")
        return

    console.print(f"[#5A9A4A]✓ 找到 {len(candidates)} 筆待拋轉銷貨單[/#5A9A4A]")

    record_ids = None
    while True:
        selected = questionary.checkbox(
            "請選擇要建立出貨單的銷貨單（空白鍵勾選，Enter 確認）：",
            choices=[questionary.Choice(c["label"], checked=False) for c in candidates],
        ).ask()
        if not selected:
            console.print("[#FF7700]返回主選單[/#FF7700]")
            return

        record_ids = [c["id"] for c in candidates if c["label"] in selected]

        console.print("[#B0A898]── 即將執行：建立出貨單 ──[/#B0A898]")
        for label in selected:
            console.print(f"  {label}")
        ok = questionary.confirm("確認執行？", default=True).ask()
        if ok:
            break

    with console.status("[#B0A898]取得按鈕設定...[/#B0A898]", spinner="dots"):
        button_id = ragic_get_action_button_id(SALES_ORDER_SHEET, "建立出貨單")
    if button_id is None:
        console.print("[red]找不到「建立出貨單」按鈕，請確認 Ragic 表單設定[/red]")
        return

    if args.dry_run:
        console.print(f"[#FF7700]★ DRY-RUN：buttonId={button_id}，對象 {record_ids}[/#FF7700]")
        return

    success = 0
    for rid in record_ids:
        try:
            result = ragic_trigger_button(SALES_ORDER_SHEET, rid, button_id)
            if result.get("status") == "SUCCESS":
                urls = result.get("urls", [])
                console.print(f"[#5A9A4A]✓ {rid} 拋轉成功[/#5A9A4A]" + (f"  → {urls[0]}" if urls else ""))
                logging.info("出貨單建立成功 sales_id=%s", rid)
                success += 1
            else:
                console.print(f"[red]✗ {rid} 拋轉失敗：{_friendly_error(result.get('msg', str(result)))}[/red]")
                logging.warning("出貨單建立失敗 sales_id=%s msg=%s", rid, result.get('msg', result))
        except Exception as e:
            console.print(f"[red]✗ {rid} 發生錯誤：{_friendly_error(str(e))}[/red]")
            logging.error("出貨單建立錯誤 sales_id=%s error=%s", rid, e)
    console.print(f"[bold #5A9A4A]完成！{success}/{len(record_ids)} 筆出貨單建立成功[/bold #5A9A4A]")
    console.print("[dim]請至 Ragic 出貨單頁面確認[/dim]")
    _pause()


def run_create_outbound_order(args):
    """出貨單拋轉建立出庫單，並自動補填子表的倉庫代碼和庫存編號。"""
    console.print("[#B0A898]── 建立出庫單（出貨單拋轉）──[/#B0A898]")
    # 載入出貨單
    with console.status("[#B0A898]載入出貨單資料...[/#B0A898]", spinner="dots"):
        records = ragic_get(DELIVERY_ORDER_SHEET)

    candidates = []
    for rid, rec in records.items():
        candidates.append({
            "id":    rid,
            "label": f"{rec.get('出貨單號','?')}  {rec.get('客戶名稱','?')}  {rec.get('訂單日期','?')}",
        })

    if not candidates:
        console.print("[#FF7700]沒有出貨單資料[/#FF7700]")
        return

    console.print(f"[#5A9A4A]✓ 找到 {len(candidates)} 筆出貨單[/#5A9A4A]")

    # 載入倉庫庫存（一次性，不隨步驟重複）
    with console.status("[#B0A898]載入倉庫庫存資料...[/#B0A898]", spinner="dots"):
        inventory = ragic_get(INVENTORY_SHEET)

    warehouses: dict = {}
    inv_by_wh_prod: Dict[tuple, list] = {}
    for rec in inventory.values():
        wh_code  = str(rec.get("倉庫代碼", "")).strip()
        wh_name  = str(rec.get("倉庫名稱", "")).strip()
        prod     = str(rec.get("商品編號", "")).strip()
        inv_code = str(rec.get("庫存編號", "")).strip()
        if wh_code:
            warehouses[wh_code] = wh_name
        if wh_code and prod and inv_code:
            inv_by_wh_prod.setdefault((wh_code, prod), []).append(inv_code)

    if not warehouses:
        console.print("[red]無法載入倉庫資料[/red]")
        return

    DEFAULT_WH = "TW01"
    BACK = "← 返回"
    sorted_wh = sorted(warehouses.items(), key=lambda x: (0 if x[0] == DEFAULT_WH else 1, x[0]))
    wh_choices = [f"{code}  {name}" for code, name in sorted_wh]
    DELIVERY_SUBTABLE = "_subtable_3000886"

    step = 1
    selected_records = record_ids = None
    warehouse_code = warehouse_name = None
    prod_inv_map = None

    while True:
        if step == 1:
            selected = questionary.checkbox(
                "請選擇要建立出庫單的出貨單（空白鍵勾選，Enter 確認）：",
                choices=[questionary.Choice(c["label"], checked=False) for c in candidates],
            ).ask()
            if not selected:
                console.print("[#FF7700]返回主選單[/#FF7700]")
                return
            selected_records = [c for c in candidates if c["label"] in selected]
            record_ids = [c["id"] for c in selected_records]
            step = 2

        elif step == 2:
            wh_sel = questionary.select("請選擇倉庫：", choices=[BACK] + wh_choices).ask()
            if not wh_sel or wh_sel == BACK:
                step = 1
                continue
            warehouse_code = wh_sel.split("  ")[0].strip()
            warehouse_name = warehouses.get(warehouse_code, "")
            step = 3

        elif step == 3:
            products_needed: list = []
            seen_prods: set = set()
            for c in selected_records:
                sub = records[c["id"]].get(DELIVERY_SUBTABLE, {})
                for row in sub.values():
                    prod = str(row.get("商品編號*", "") or row.get("商品編號", "")).strip()
                    if prod and prod not in seen_prods:
                        seen_prods.add(prod)
                        products_needed.append({"prod": prod, "name": str(row.get("商品名稱", "")).strip()})

            prod_inv_map = {}
            cancelled = False
            for item in products_needed:
                prod = item["prod"]
                options = inv_by_wh_prod.get((warehouse_code, prod), [])
                if not options:
                    console.print(f"[#FF7700]⚠ {prod} 在 {warehouse_code} 無庫存紀錄，跳過[/#FF7700]")
                    prod_inv_map[prod] = ""
                    continue
                if len(options) == 1:
                    prod_inv_map[prod] = options[0]
                    console.print(f"[dim]{prod} {item['name']} → {options[0]}（唯一選項，自動帶入）[/dim]")
                else:
                    inv_sel = questionary.select(
                        f"請選擇 {prod} {item['name']} 的庫存編號：",
                        choices=[BACK] + options,
                    ).ask()
                    if not inv_sel or inv_sel == BACK:
                        cancelled = True
                        break
                    prod_inv_map[prod] = inv_sel
            if cancelled:
                step = 2
                continue
            step = 4

        elif step == 4:
            console.print("[#B0A898]── 即將執行：建立出庫單 ──[/#B0A898]")
            for c in selected_records:
                console.print(f"  {c['label']}")
            console.print(f"  倉庫：{warehouse_code}  {warehouse_name}")
            for prod, inv in prod_inv_map.items():
                if inv:
                    console.print(f"  {prod} → {inv}")
            console.print("[dim]  拋轉後自動補：倉庫/庫存編號、單據備註=客戶名稱[/PO#]、明細備註=【EAN】條碼[/dim]")
            console.print("[dim]  並偵測出貨拆盒：整中盒會再請你確認後才改庫存、零頭只提醒拆實體[/dim]")
            ok = questionary.confirm("確認執行？", default=True).ask()
            if not ok:
                step = 1
                continue
            break

    # ── 自動拆盒（中盒→單盒）：拋轉「之前」先補足單盒，否則出庫扣單盒會不足 ──
    # 逐出貨單明細分開判（同商品跨單不合併）。整中盒→自動拆、散單盒不動；
    # 非整中盒→零售/填錯只提醒拆實體；中盒線不拆。多單拆同中盒會累計檢查。
    line_needs = []
    for c in selected_records:
        rec = records[c["id"]]
        label = f"{rec.get('出貨單號', c['id'])} {rec.get('客戶名稱', '')}".strip()
        agg = {}
        for row in rec.get(DELIVERY_SUBTABLE, {}).values():
            prod = str(row.get("商品編號*", "") or row.get("商品編號", "")).strip()
            if prod:
                agg[prod] = agg.get(prod, 0) + _to_int(row.get("數量", 0), 0)
        for prod, q in agg.items():
            line_needs.append((label, prod, q))
    break_plan = compute_break_plan(line_needs, inventory, warehouse_code)
    auto   = [p for p in break_plan if p["status"] == "ok"]
    manual = [p for p in break_plan if p["status"] == "manual" and max(0, p["need"] - p["have"]) > 0]
    issues = [p for p in break_plan if p["status"] in ("parent_short", "no_parent", "no_stock")]
    if selected_records:
        console.print("[#B0A898]── 出貨拆盒（客戶下 pcs、實出中盒，拆中盒扣帳）──[/#B0A898]")
        # 每張選到的單都列：要拆的列明細，不用拆的標「✓ 無需拆盒」讓人安心
        all_labels = []
        for c in selected_records:
            rec = records[c["id"]]
            lab = f"{rec.get('出貨單號', c['id'])} {rec.get('客戶名稱', '')}".strip()
            if lab not in all_labels:
                all_labels.append(lab)
        for lab in all_labels:
            a = [x for x in auto if x["label"] == lab]
            m = [x for x in manual if x["label"] == lab]
            i = [x for x in issues if x["label"] == lab]
            console.print(f"  [bold]【{lab}】[/bold]")
            if not (a or m or i):
                console.print("     [#5A9A4A]✓ 無需拆盒[/#5A9A4A]")
                continue
            for p in a:
                console.print(f"     [#5A9A4A]{p['prod']}[/#5A9A4A] {p['name'][:9]}  客戶 {p['need']} → 拆 {p['parent']} 中盒 ×{p['boxes']}"
                              f"（中盒 {p['parent_qty']}→{p['parent_qty'] - p['boxes']}）")
            for p in m:
                console.print(f"     [#FF7700]⚠ {p['prod']} {p['name'][:9]} 客戶 {p['need']}（非整中盒）"
                              f"→ 用散盒，不足請拆實體 {p['boxes']} 盒（不自動）[/#FF7700]")
            for p in i:
                if p["status"] == "parent_short":
                    console.print(f"     [red]⛔ {p['prod']} 客戶{p['need']}，中盒{p['parent']}剩{p.get('avail', p.get('parent_qty'))}不夠（需{p['boxes']}）→ 請先補中盒[/red]")
                else:
                    console.print(f"     [red]⛔ {p['prod']} 客戶{p['need']}，查無庫存或無中盒可拆 → 請人工處理[/red]")
        # 同一中盒跨多單只 PATCH 一次（中盒扣總盒數、單盒加總）
        merged = {}
        for p in auto:
            m = merged.setdefault(p["parent"], {
                "parent_rid": p["parent_rid"], "parent_qty": p["parent_qty"],
                "unit_rid": p["unit_rid"], "unit_qty": p["unit_qty"],
                "boxes": 0, "gain": 0, "unit": p["prod"]})
            m["boxes"] += p["boxes"]
            m["gain"] += p["gain"]
        if args.dry_run:
            console.print("[#FF7700]★ DRY-RUN：僅預覽拆盒，未改任何庫存[/#FF7700]")
        elif merged:
            if questionary.confirm(f"確認拆盒（{len(merged)} 種中盒、共 {sum(m['boxes'] for m in merged.values())} 盒）？", default=True).ask():
                for pc, m in merged.items():
                    try:
                        ragic_patch(INVENTORY_SHEET, m["parent_rid"], {INVENTORY_QTY_CID: m["parent_qty"] - m["boxes"]})
                        ragic_patch(INVENTORY_SHEET, m["unit_rid"],   {INVENTORY_QTY_CID: m["unit_qty"] + m["gain"]})
                        console.print(f"[#5A9A4A]✓ 拆盒 {pc} 中盒 -{m['boxes']}、{m['unit']} 單盒 +{m['gain']}[/#5A9A4A]")
                        logging.info("拆盒 wh=%s 中盒=%s boxes=%s gain=%s", warehouse_code, pc, m["boxes"], m["gain"])
                    except Exception as e:
                        console.print(f"[red]⚠ 拆盒 {pc} 失敗：{_friendly_error(str(e))}[/red]")
                        logging.error("拆盒失敗 中盒=%s error=%s", pc, e)
            else:
                console.print("[#FF7700]略過自動拆盒（你可手動處理後再繼續）[/#FF7700]")

    with console.status("[#B0A898]取得按鈕設定...[/#B0A898]", spinner="dots"):
        button_id = ragic_get_action_button_id(DELIVERY_ORDER_SHEET, "建立出庫單")
    if button_id is None:
        console.print("[red]找不到「建立出庫單」按鈕，請確認 Ragic 表單設定[/red]")
        return

    if args.dry_run:
        console.print(f"[#FF7700]★ DRY-RUN：buttonId={button_id}，倉庫={warehouse_code}，對象 {record_ids}[/#FF7700]")
        return

    # 記錄觸發前的出庫單 ID
    console.print("[#B0A898]記錄現有出庫單...[/#B0A898]")
    before_ids = set(ragic_get(OUTBOUND_ORDER_SHEET).keys())

    console.print(f"[#B0A898]逐筆觸發建立出庫單（{len(record_ids)} 筆）...[/#B0A898]")
    for rid in record_ids:
        try:
            result = ragic_trigger_button(DELIVERY_ORDER_SHEET, rid, button_id)
            if result.get("status") == "SUCCESS":
                console.print(f"[#5A9A4A]✓ {rid} 拋轉成功[/#5A9A4A]")
                logging.info("出庫單觸發成功 delivery_id=%s", rid)
            else:
                console.print(f"[red]✗ {rid} 拋轉失敗：{_friendly_error(result.get('msg', str(result)))}[/red]")
                logging.warning("出庫單觸發失敗 delivery_id=%s msg=%s", rid, result.get('msg', result))
        except Exception as e:
            console.print(f"[red]✗ {rid} 發生錯誤：{_friendly_error(str(e))}[/red]")
            logging.error("出庫單觸發錯誤 delivery_id=%s error=%s", rid, e)

    console.print("[dim]等待 Ragic 建立出庫單（3 秒）...[/dim]")
    time.sleep(3)

    after_records = ragic_get(OUTBOUND_ORDER_SHEET)
    new_ids = set(after_records.keys()) - before_ids
    if not new_ids:
        console.print("[#FF7700]⚠ 未偵測到新建立的出庫單（可能已被 Ragic 擋掉重複拋轉）[/#FF7700]")
        return

    console.print(f"[#5A9A4A]✓ 偵測到 {len(new_ids)} 筆新出庫單，開始補填倉庫資料...[/#5A9A4A]")

    # ── 備註自動填寫準備 ───────────────────────────────────────
    # 單據備註(表頭)=客戶名稱 [+ PO#]；子表每列備註=國際條碼（一律，不分客戶）。
    # 客戶名稱/訂單編號從來源出貨單帶（出庫單表頭沒有），用出貨單號對回。
    shipno_to_cust = {
        str(records[rid].get("出貨單號", "")).strip(): str(records[rid].get("客戶名稱", "")).strip()
        for rid in record_ids
    }
    # PO# 來自銷貨單備註：出貨單.訂單編號 → 銷貨單.訂單編號 → 備註裡的 PO#
    shipno_to_order = {
        str(records[rid].get("出貨單號", "")).strip(): str(records[rid].get("訂單編號", "")).strip()
        for rid in record_ids
    }
    order_to_po = {}
    try:
        for so in ragic_get(SALES_ORDER_SHEET).values():
            on = str(so.get("訂單編號", "")).strip()
            po = extract_po(so.get("備註", ""))
            if on and po:
                order_to_po[on] = po
    except Exception as e:
        logging.warning("載入銷貨單帶 PO# 失敗（單據備註將只填客戶名稱）：%s", e)
    code_to_barcode = build_code_to_barcode(load_price_index())

    patched = 0
    for oid in new_ids:
        rec = after_records[oid]
        subtable = rec.get(OUTBOUND_ITEMS_SUBTABLE_KEY, {})
        if not subtable:
            console.print(f"[#FF7700]⚠ 出庫單 {oid} 沒有子表項目，跳過[/#FF7700]")
            continue

        ship_no = str(rec.get("出貨單號", "")).strip()
        customer = shipno_to_cust.get(ship_no, "")
        po = order_to_po.get(shipno_to_order.get(ship_no, ""), "")
        doc_note = f"{customer} / PO#{po}".strip(" /") if po else customer

        # 填倉庫代碼、庫存編號（用 CID，必填欄位用欄位名稱會被 Ragic validation 擋掉）＋
        # 子表備註=國際條碼。倉庫名稱(3001125)唯讀，填代碼後 Ragic 自動帶入。
        updated_rows = {}
        for row_id, row in subtable.items():
            if str(row_id).startswith("_"):
                continue
            prod = str(row.get("商品編號", "")).strip()
            cell = {}
            inv_code = prod_inv_map.get(prod, "")
            if inv_code:
                cell["3001124"] = warehouse_code  # 倉庫代碼
                cell["3001126"] = inv_code         # 庫存編號
            else:
                console.print(f"[#FF7700]⚠ 出庫單 {oid} 商品 {prod} 無庫存編號，該列倉庫欄位略過[/#FF7700]")
                logging.warning("出庫單 %s 商品 %s 無庫存編號，略過", oid, prod)
            bars = code_to_barcode.get(prod, [])
            if bars:
                cell[OUTBOUND_ROW_NOTE_CID] = f"【EAN】{bars[0]}"  # 國際條碼
                if len(bars) > 1:
                    console.print(f"[#FF7700]⚠ {prod} 有多個國際條碼 {bars}，備註取 {bars[0]}[/#FF7700]")
            else:
                console.print(f"[#FF7700]⚠ {prod} 查無國際條碼，該列備註留空[/#FF7700]")
            if cell:
                updated_rows[str(row_id)] = cell

        patch_body = {OUTBOUND_ITEMS_SUBTABLE_KEY: updated_rows}
        if doc_note:
            patch_body[OUTBOUND_DOC_NOTE_CID] = doc_note  # 單據備註=客戶名稱[ / PO#]

        try:
            ragic_patch(OUTBOUND_ORDER_SHEET, oid, patch_body)
            patched += 1
            console.print(f"[#5A9A4A]✓ 出庫單 {oid} 補填完成（{warehouse_code}／單據備註「{doc_note or '-'}」＋備註國際條碼）[/#5A9A4A]")
            logging.info("出庫單補填成功 outbound_id=%s warehouse=%s doc_note=%s", oid, warehouse_code, doc_note)
        except Exception as e:
            console.print(f"[red]⚠ 出庫單 {oid} 補填失敗：{e}[/red]")
            logging.error("出庫單補填失敗 outbound_id=%s error=%s", oid, e)

    console.print(f"[bold #5A9A4A]完成！{patched}/{len(new_ids)} 筆出庫單已補填倉庫資料[/bold #5A9A4A]")
    console.print("[dim]請至 Ragic 出庫單頁面確認[/dim]")
    _pause()


def run_export_inventory(args, price_index: dict):
    """從 Ragic 倉庫庫存匯出 Excel，自動換算 PCS 填入客戶模板的現貨欄位。"""
    import copy
    import openpyxl

    BASE_OUTPUT.mkdir(exist_ok=True)

    # ── 倉庫選擇 ─────────────────────────────────────────────
    with console.status("[#B0A898]載入倉庫庫存資料...[/#B0A898]", spinner="dots"):
        inventory_all = ragic_get(INVENTORY_SHEET)

    warehouses: dict = {}
    for rec in inventory_all.values():
        wh_code = str(rec.get("倉庫代碼", "")).strip()
        wh_name = str(rec.get("倉庫名稱", "")).strip()
        if wh_code:
            warehouses[wh_code] = wh_name

    if not warehouses:
        console.print("[red]無法載入倉庫資料[/red]")
        return

    DEFAULT_WH = "TW01"
    BACK = "← 返回"
    sorted_wh = sorted(warehouses.items(), key=lambda x: (0 if x[0] == DEFAULT_WH else 1, x[0]))
    wh_choices = [f"{code}  {name}" for code, name in sorted_wh]

    wh_sel = questionary.select("請選擇倉庫：", choices=[BACK] + wh_choices).ask()
    if not wh_sel or wh_sel == BACK:
        console.print("[#FF7700]返回主選單[/#FF7700]")
        return
    warehouse_code = wh_sel.split("  ")[0].strip()
    warehouse_name = warehouses.get(warehouse_code, "")

    # ── 模板選擇 ─────────────────────────────────────────────
    BASE_TEMPLATES.mkdir(exist_ok=True)
    templates = sorted(BASE_TEMPLATES.glob("*.xlsx"), reverse=True)
    if not templates:
        console.print(f"[red]找不到模板，請將 .xlsx 模板放入 {BASE_TEMPLATES}[/red]")
        return

    TPL_DISPLAY = {
        "quote-template.xlsx":     "quote-template.xlsx（報價單）",
        "inventory-template.xlsx": "inventory-template.xlsx（庫存總覽）",
    }
    tpl_map = {TPL_DISPLAY.get(t.name, t.name): t for t in templates}
    selected = questionary.checkbox(
        "請選擇模板（空白鍵勾選，Enter 確認；不選直接 Enter 返回）：",
        choices=[questionary.Choice(label, checked=False) for label in tpl_map],
    ).ask()
    if not selected:
        console.print("[#FF7700]返回主選單[/#FF7700]")
        return

    tpl_path = tpl_map[selected[0]]

    # ── 確認 ─────────────────────────────────────────────────
    console.print(f"[#B0A898]── 即將執行：匯出庫存報表 ──[/#B0A898]")
    console.print(f"  倉庫：{warehouse_code}  {warehouse_name}")
    console.print(f"  模板：{tpl_path.name}")
    ok = questionary.confirm("確認執行？", default=True).ask()
    if not ok:
        return

    # ── 建立 product_code → barcode 反向索引 ─────────────────
    # 商品單價代號格式為 BBB042-1（有尾綴），庫存商品編號為 BBB042（無尾綴）
    # 去掉 -數字 尾綴後建立反向索引
    import re as _re
    code_to_barcode: Dict[str, str] = {}
    for barcode, entries in price_index.items():
        for entry in entries:
            base = _re.sub(r'-\d+$', '', entry["product_code"])
            code_to_barcode[base] = barcode

    # ── 計算各條碼的 PCS（只算 spec > 1 的）────────────────
    # 報客戶的單位（中盒/箱類），其餘（單盒/個/袋等）跳過
    BULK_UNITS = {"中盒", "箱", "整箱", "端盒"}

    inventory_pcs: Dict[str, int] = {}
    skipped_single = 0
    for rec in inventory_all.values():
        if str(rec.get("倉庫代碼", "")).strip() != warehouse_code:
            continue
        unit = str(rec.get("單位", "")).strip()
        if unit not in BULK_UNITS:
            skipped_single += 1
            continue
        prod_code = str(rec.get("商品編號", "")).strip()
        qty_raw = rec.get("數量", 0)
        spec_raw = rec.get("規格", "1")
        try:
            qty = int(float(qty_raw or 0))
        except (ValueError, TypeError):
            qty = 0
        try:
            spec = int(float(spec_raw or 1))
        except (ValueError, TypeError):
            spec = 1

        barcode = code_to_barcode.get(prod_code)
        if not barcode:
            continue

        pcs = qty * spec
        inventory_pcs[barcode] = inventory_pcs.get(barcode, 0) + pcs

    console.print(f"[#5A9A4A]✓ 計算完成：{len(inventory_pcs)} 種條碼有庫存（略過 {skipped_single} 筆單盒項目）[/#5A9A4A]")

    # ── 填入模板 ─────────────────────────────────────────────
    wb = openpyxl.load_workbook(tpl_path)
    ws = wb.active

    # 自動偵測現貨欄位置（從 row 2 或 row 3 找「現貨」）
    inv_col_idx = None
    for check_row in (2, 3):
        for cell in ws[check_row]:
            if str(cell.value or '').strip() == '現貨':
                inv_col_idx = cell.column - 1  # 轉為 0-indexed
                break
        if inv_col_idx is not None:
            break
    if inv_col_idx is None:
        console.print("[red]✗ 此模板找不到「現貨」欄位，無法匯出庫存。請選擇 inventory 或 quote 模板。[/red]")
        return

    filled = 0
    for row in ws.iter_rows(min_row=4):
        d_cell = row[3]  # D 欄 index=3
        if d_cell.value is None:
            continue
        try:
            barcode = str(int(float(d_cell.value)))
        except (ValueError, TypeError):
            continue
        if barcode in inventory_pcs and inv_col_idx < len(row):
            row[inv_col_idx].value = inventory_pcs[barcode]
            filled += 1

    # ── 強制 Excel 開檔重算公式（避免快取值顯示為 0）──
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    # ── 儲存輸出 ─────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    tpl_prefix = tpl_path.stem.replace("-template", "")
    out_path = BASE_OUTPUT / f"{tpl_prefix}_{warehouse_code}_{ts}.xlsx"
    wb.save(out_path)

    # ── 保留範本的嵌入圖片（openpyxl 3.x 不支援 oneCellAnchor，存檔會掉圖）──
    # 從範本 zip 注入 media/drawings/charts/embeddings 與相關 rels & content types，
    # 只用 openpyxl 輸出檔取代邏輯內容（sheet/sharedStrings/styles/workbook）。
    import zipfile
    with zipfile.ZipFile(tpl_path) as ztpl:
        merged = {n: ztpl.read(n) for n in ztpl.namelist()}
    with zipfile.ZipFile(out_path) as zout:
        oxl = {n: zout.read(n) for n in zout.namelist()}
    for f in ('xl/worksheets/sheet1.xml', 'xl/sharedStrings.xml',
              'xl/styles.xml', 'xl/workbook.xml'):
        if f in oxl:
            merged[f] = oxl[f]
    merged.pop('xl/calcChain.xml', None)  # Excel 會自動重建
    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zfinal:
        for name, data in merged.items():
            zfinal.writestr(name, data)

    console.print(f"[bold #5A9A4A]✓ 完成！填入 {filled} 筆，輸出至：{out_path}[/bold #5A9A4A]")
    logging.info("庫存報表匯出成功 warehouse=%s filled=%d path=%s", warehouse_code, filled, out_path)
    _pause()


# ── 歡迎畫面 ─────────────────────────────────────────────────

def _get_current_user() -> str:
    """取得用戶名稱：優先 Ragic API，失敗則用系統登入名。"""
    try:
        if _KEY_FILE.exists():
            api_key = _KEY_FILE.read_text(encoding="utf-8").strip()
            url = f"{RAGIC_BASE}/{RAGIC_ACCOUNT}?api&getUserInfo=true"
            resp = requests.get(url, headers={"Authorization": f"Basic {api_key}"}, timeout=3)
            data = resp.json()
            name = (data.get("name") or data.get("fullName") or
                    data.get("userName") or data.get("user", {}).get("name", ""))
            if name:
                return name
    except Exception:
        pass
    try:
        import getpass
        return getpass.getuser()
    except Exception:
        return ""


def _calc_revenue(data: dict, date_from: str, date_to: str) -> float:
    """加總指定日期範圍內銷貨單總計。"""
    total = 0.0
    for rec in data.values():
        order_date = rec.get("訂單日期", rec.get("日期", ""))
        if not order_date:
            continue
        d = order_date[:10].replace("-", "/")
        if date_from <= d <= date_to:
            val = rec.get("總金額(含稅)", rec.get("小計", rec.get("總計", "0"))) or "0"
            try:
                total += float(str(val).replace(",", ""))
            except ValueError:
                pass
    return total


def _compute_revenue_summaries(data: dict, username: str = "") -> tuple:
    """
    從已抓取的銷貨單資料計算公司營業額與個人業績。
    回傳 (company_rows, personal_stats)
      company_rows : [(label, amount_str), ...]
      personal_stats: {"year": float, "count": int} 或 None
    """
    import datetime as dt
    try:
        today = date.today()

        first_this_month = today.replace(day=1)
        lm_end = first_this_month - dt.timedelta(days=1)
        lm_start = lm_end.replace(day=1)
        q = (today.month - 1) // 3
        q_start = date(today.year, q * 3 + 1, 1)
        y_start = date(today.year, 1, 1)

        company_rows = []
        lm_total = _calc_revenue(data, lm_start.strftime("%Y/%m/%d"), lm_end.strftime("%Y/%m/%d"))
        if lm_total:
            company_rows.append((f"上月 ({lm_start.strftime('%Y/%m')})", f"NT$ {lm_total:,.0f}"))
        q_total = _calc_revenue(data, q_start.strftime("%Y/%m/%d"), today.strftime("%Y/%m/%d"))
        if q_total:
            company_rows.append((f"本季 (Q{q + 1})", f"NT$ {q_total:,.0f}"))
        y_total = _calc_revenue(data, y_start.strftime("%Y/%m/%d"), today.strftime("%Y/%m/%d"))
        if y_total:
            company_rows.append((f"本年 ({today.year})", f"NT$ {y_total:,.0f}"))

        personal_stats = None
        if username:
            py_total = 0.0
            py_count = 0
            y_from = y_start.strftime("%Y/%m/%d")
            y_to = today.strftime("%Y/%m/%d")
            for rec in data.values():
                if not isinstance(rec, dict): continue
                if rec.get("建檔人員", "").lower() != username.lower(): continue
                order_date = rec.get("訂單日期", "")
                if not order_date: continue
                d = order_date[:10].replace("-", "/")
                if y_from <= d <= y_to:
                    val = rec.get("總金額(含稅)", rec.get("小計", "0")) or "0"
                    try:
                        py_total += float(str(val).replace(",", ""))
                        py_count += 1
                    except ValueError:
                        pass
            if py_count:
                personal_stats = {"year": py_total, "count": py_count}

        return company_rows, personal_stats
    except Exception:
        return [], None


def _get_revenue_summary() -> list:
    """相容舊呼叫，回傳公司營業額列表。"""
    try:
        data = ragic_get(SALES_ORDER_SHEET, limit=2000)
        rows, _ = _compute_revenue_summaries(data)
        return rows
    except Exception:
        return []


def _get_recent_activity() -> list:
    """從 upload_log.json 讀取最近操作，回傳 [(日期, 描述), ...] 最多 5 筆。"""
    from collections import defaultdict
    log = _load_upload_log()
    if not log:
        return []
    date_counts: dict = defaultdict(int)
    for v in log.values():
        date = v.get("uploaded_at", "")[:10]
        if date:
            date_counts[date] += 1
    sorted_dates = sorted(date_counts.items(), key=lambda x: x[0], reverse=True)[:5]
    return [(d, f"銷貨單 × {c} 筆") for d, c in sorted_dates]


def _show_welcome():
    """顯示仿 Claude Code 風格的歡迎畫面。兩支 API 並行抓取，加速啟動。"""
    import threading
    results = {"username": "", "sales_data": {}, "activity": _get_recent_activity()}

    def _fetch_user():
        results["username"] = _get_current_user()

    def _fetch_sales():
        try:
            results["sales_data"] = ragic_get(SALES_ORDER_SHEET, limit=2000)
        except Exception:
            results["sales_data"] = {}

    t1 = threading.Thread(target=_fetch_user, daemon=True)
    t2 = threading.Thread(target=_fetch_sales, daemon=True)
    t1.start(); t2.start()
    t1.join(timeout=4); t2.join(timeout=4)

    username = results["username"]
    welcome_line = f"歡迎回來，{username.capitalize()}！" if username else "歡迎回來！"
    revenue_rows, personal_stats = _compute_revenue_summaries(results["sales_data"], username)

    # 左欄
    left = Table.grid(padding=(0, 2))
    left.add_column()
    left.add_row(Text(welcome_line, style="bold #D4C9B0"))
    if personal_stats:
        today_year = date.today().year
        p_table = Table.grid(padding=(0, 1))
        p_table.add_column(no_wrap=True)
        p_table.add_column(style="bold #B8860B", no_wrap=True)
        p_table.add_column(style="#B0A898", no_wrap=True)
        p_table.add_row(
            Text(f"我的業績（{today_year}）", style="#B0A898"),
            f"NT$ {personal_stats['year']:,.0f}",
            f"  · {personal_stats['count']} 筆",
        )
        left.add_row(p_table)
    left.add_row("")
    left.add_row(Text("Boptoys", style="bold #C5A059"))
    left.add_row(Text("潮玩波普國際有限公司", style="#C5A059"))
    left.add_row(Text("統一編號 82906411", style="dim"))
    if revenue_rows:
        left.add_row("")
        rev_table = Table.grid(padding=(0, 2))
        rev_table.add_column(style="dim", no_wrap=True)
        rev_table.add_column(style="bold #FF7700", no_wrap=True)
        for label, amount in revenue_rows:
            rev_table.add_row(label, amount)
        left.add_row(rev_table)

    # 右欄：最近操作
    activity = _get_recent_activity()
    right = Table.grid(padding=(0, 2))
    right.add_column(style="dim", no_wrap=True)
    right.add_column()
    if activity:
        for d, desc in activity:
            right.add_row(d, desc)
    else:
        right.add_row("尚無操作紀錄", "")

    # 組合成雙欄
    layout = Table.grid(expand=True, padding=(0, 1))
    layout.add_column(ratio=3)
    layout.add_column(ratio=2)
    layout.add_row(left, right)

    console.print(Panel(layout, title="[bold #C5A059]Ragic ERP Tools[/bold #C5A059]", subtitle=f"[dim]v{APP_VERSION}[/dim]", border_style="#C5A059"))
    console.print(Rule(style="#C5A059"))


# ── 批次發樣／開免費單 ─────────────────────────────────────────

def _load_json_file(path: str) -> dict:
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_json_file(path: str, data: dict) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _load_sample_products() -> list:
    """載入商品單價 20006 全品項供樣品搜尋。直接讀全表（不靠國際條碼過濾），
    這樣展示架等「無條碼」商品也搜得到（load_price_index 會略過無條碼者）。"""
    with console.status("[#B0A898]載入商品清單...[/#B0A898]", spinner="dots"):
        records = ragic_get(PRODUCT_PRICE_SHEET)
    out = []
    for rec in records.values():
        code = str(rec.get("商品單價代號", "")).strip()
        if not code:
            continue
        out.append({
            "code":    code,
            "name":    str(rec.get("商品名稱", "")),
            "spec":    rec.get("規格", 1),
            "unit":    str(rec.get("單位", "")),
            "barcode": str(rec.get("國際條碼", "")).strip(),
        })
    console.print(f"[#5A9A4A]✓ 載入 {len(out)} 項商品[/#5A9A4A]")
    return out


def _search_products(products: list, keyword: str) -> list:
    """模糊比對：商品代號／名稱／條碼任一含關鍵字（大小寫不分、中英皆可）。"""
    kw = keyword.strip().lower()
    if not kw:
        return []
    return [p for p in products
            if kw in p["code"].lower() or kw in p["name"].lower() or kw in p["barcode"].lower()]


def _edit_combo_items(products: list, items: list) -> Optional[list]:
    """組合編輯器（可加可減）。傳入起始 items（新建給 []，編輯給現有），回最終 items 或 None 取消。"""
    BACK = "← 返回"
    ADD, REMOVE, DONE, CANCEL = "➕ 加品項", "➖ 移除品項", "✓ 完成", "✗ 取消"
    items = [dict(it) for it in items]  # 複製，不動原本
    while True:
        if items:
            console.print("[#B0A898]目前組合：[/#B0A898]")
            for it in items:
                console.print(f"  {it['code']:<12} {it['name'][:20]} ×{it['qty']}")
        else:
            console.print("[dim]目前組合：（空）[/dim]")
        opts = [ADD] + ([REMOVE, DONE] if items else []) + [CANCEL]
        act = _select_with_esc("組合編輯：", choices=opts)
        if not act or act == CANCEL:
            return None
        if act == DONE:
            return items
        if act == REMOVE:
            ropts = [f"{it['code']} {it['name'][:16]} ×{it['qty']}" for it in items] + [BACK]
            rsel = _select_with_esc("移除哪一項？", choices=ropts)
            if rsel and rsel != BACK:
                removed = items.pop(ropts.index(rsel))
                console.print(f"[#5A9A4A]✓ 已移除 {removed['code']}[/#5A9A4A]")
            continue
        # 加品項：搜尋
        kw = questionary.text("搜尋商品（中文/英文/代號；留空取消）：", default="").ask()
        if not kw or not kw.strip():
            continue
        hits = _search_products(products, kw.strip())
        if not hits:
            console.print("[#FF7700]查無商品，換個關鍵字[/#FF7700]")
            continue
        hopts = [f"{h['code']:<12} {h['name'][:24]}（{h['unit']}）" for h in hits[:25]] + [BACK]
        psel = _select_with_esc(f"找到 {len(hits)} 筆，請選：", choices=hopts)
        if not psel or psel == BACK:
            continue
        prod = hits[hopts.index(psel)]
        qraw = questionary.text(f"{prod['name'][:16]} 數量（pcs）：", default="1").ask()
        try:
            qn = int(float(qraw))
        except (TypeError, ValueError):
            console.print("[#FF7700]數量需為數字，略過[/#FF7700]")
            continue
        if qn <= 0:
            continue
        exist = next((it for it in items if it["code"] == prod["code"]), None)
        if exist:
            exist["qty"] = qn   # 同品項 → 更新數量
            console.print(f"[#5A9A4A]✓ 已更新 {prod['code']} ×{qn}[/#5A9A4A]")
        else:
            items.append({"code": prod["code"], "name": prod["name"], "qty": qn})
            console.print(f"[#5A9A4A]✓ 已加入 {prod['code']} ×{qn}[/#5A9A4A]")


def _pick_sample_combo(products: list) -> Optional[list]:
    """選／新建／編輯／刪除組合範本。回 [{code,name,qty}] 或 None（取消）。"""
    BACK = "← 返回"
    NEW, EDIT, DELETE = "【新建組合】", "✏️ 編輯組合", "🗑 刪除組合"
    combos = _load_json_file(SAMPLE_COMBOS_FILE)
    while True:
        choices = [NEW]
        for name, items in combos.items():
            summary = "、".join("{}×{}".format(it["name"][:8], it["qty"]) for it in items)
            choices.append("{}（{}）".format(name, summary))
        if combos:
            choices += [EDIT, DELETE]
        choices.append(BACK)
        sel = _select_with_esc("請選擇樣品組合：", choices=choices)
        if not sel or sel == BACK:
            return None
        if sel == DELETE:
            dsel = _select_with_esc("要刪除哪個組合？", choices=list(combos.keys()) + [BACK])
            if dsel and dsel != BACK and questionary.confirm(f"確定刪除組合「{dsel}」？", default=False).ask():
                combos.pop(dsel, None)
                _save_json_file(SAMPLE_COMBOS_FILE, combos)
                console.print(f"[#5A9A4A]✓ 已刪除組合「{dsel}」[/#5A9A4A]")
            continue
        if sel == EDIT:
            esel = _select_with_esc("要編輯哪個組合？", choices=list(combos.keys()) + [BACK])
            if not esel or esel == BACK:
                continue
            edited = _edit_combo_items(products, combos[esel])
            if edited is None:
                continue
            if not edited:
                console.print("[#FF7700]組合空了，未儲存（要刪請用刪除組合）[/#FF7700]")
                continue
            combos[esel] = edited
            _save_json_file(SAMPLE_COMBOS_FILE, combos)
            console.print(f"[#5A9A4A]✓ 已更新組合「{esel}」[/#5A9A4A]")
            continue
        if sel != NEW:
            cname = sel.split("（")[0]
            return combos[cname]
        # 新建組合
        items = _edit_combo_items(products, [])
        if not items:
            continue          # 取消 → 回組合選單，不退出整個功能
        name = questionary.text("為這個組合命名（留空＝這次用、不存檔）：", default="").ask()
        if name and name.strip():
            combos[name.strip()] = items
            _save_json_file(SAMPLE_COMBOS_FILE, combos)
            console.print(f"[#5A9A4A]✓ 已存組合「{name.strip()}」，下次可直接套[/#5A9A4A]")
        return items


def _pick_sample_customers(customers: list) -> Optional[list]:
    """選客戶：套固定名單打底 + 當次複選加減；可存新名單。回 [客戶dict] 或 None。"""
    BACK = "← 返回"
    DELETE = "🗑 刪除名單"
    lists = _load_json_file(SAMPLE_CUSTLIST_FILE)
    by_code = {c["code"]: c for c in customers if c["code"]}
    preset_codes = []
    applied_name = ""   # 套用了哪個名單（存檔時預設帶同名＝編輯覆蓋）
    while lists:
        opts = ["【不套，全部手選】"] + list(lists.keys()) + [DELETE, BACK]
        sel = _select_with_esc("套用固定發樣名單？（套用後可加減客戶、存回同名＝編輯）", choices=opts)
        if not sel:
            return None
        if sel == DELETE:
            dnames = list(lists.keys()) + [BACK]
            dsel = _select_with_esc("要刪除哪個名單？", choices=dnames)
            if dsel and dsel != BACK:
                if questionary.confirm(f"確定刪除名單「{dsel}」？", default=False).ask():
                    lists.pop(dsel, None)
                    _save_json_file(SAMPLE_CUSTLIST_FILE, lists)
                    console.print(f"[#5A9A4A]✓ 已刪除名單「{dsel}」[/#5A9A4A]")
            continue
        if sel != "【不套，全部手選】" and sel != BACK:
            preset_codes = [c for c in lists[sel] if c in by_code]
            applied_name = sel
        break
    # 客戶複選：questionary 的內建搜尋只吃 ASCII（中文進不去），故改「文字框打關鍵字
    # →篩小清單→勾選」，可多次搜尋累加。文字框吃得了中文，自己做子字串比對。
    all_custs = [c for c in customers if c["code"]]
    chosen_codes = set(preset_codes)
    if preset_codes:
        console.print(f"[dim]  已套用名單，預選 {len(preset_codes)} 個客戶；可再搜尋加減[/dim]")
    while True:
        tip = f"（已選 {len(chosen_codes)} 個）" if chosen_codes else ""
        kw = questionary.text(
            f"打關鍵字搜尋客戶加選{tip}；想換關鍵字就再打一次；直接 Enter＝完成全部選擇：",
            default="").ask()
        if kw is None or not kw.strip():
            if chosen_codes:
                break
            if questionary.confirm("尚未選任何客戶，放棄發樣？", default=False).ask():
                return None
            continue
        k = kw.strip().lower()
        subset = [c for c in all_custs if k in c["name"].lower() or k in c["code"].lower()]
        if not subset:
            console.print("[#FF7700]查無客戶，換個關鍵字[/#FF7700]")
            continue
        labels = [f"{c['name']}｜{c['code']}" for c in subset]
        picked = questionary.checkbox(
            f"找到 {len(subset)} 筆：空白鍵勾選 → 按 Enter 回搜尋框（可換關鍵字再搜、或在搜尋框直接 Enter 結束）：",
            choices=[questionary.Choice(labels[i], checked=(subset[i]["code"] in chosen_codes))
                     for i in range(len(subset))],
        ).ask()
        if picked is None:
            continue
        picked_set = set(picked)
        for i, c in enumerate(subset):   # 這批內：勾的加入、沒勾的移除（可在子集內取消）
            if labels[i] in picked_set:
                chosen_codes.add(c["code"])
            else:
                chosen_codes.discard(c["code"])
        console.print(f"[#5A9A4A]目前已選 {len(chosen_codes)} 個客戶[/#5A9A4A]")
    chosen = [by_code[cc] for cc in chosen_codes if cc in by_code]
    if not chosen:
        return None
    # 存成名單：套用過名單時預設帶同名（Enter 覆蓋＝編輯）；新名單則打新名字
    prompt = ("存回名單（Enter＝更新「%s」、改名＝另存、清空＝不存）：" % applied_name
              if applied_name else "把這份客戶存成固定名單？（輸入名稱＝存、留空＝不存）：")
    save = questionary.text(prompt, default=applied_name).ask()
    if save and save.strip():
        lists[save.strip()] = [c["code"] for c in chosen]
        _save_json_file(SAMPLE_CUSTLIST_FILE, lists)
        console.print(f"[#5A9A4A]✓ 已存名單「{save.strip()}」（{len(chosen)} 客戶）[/#5A9A4A]")
    return chosen


def run_sample_orders(args):
    """批次發樣／開免費單：一個組合 → 複製發給 N 個客戶，開單別=樣品/公關/贈品、單價全0、狀態未出貨。"""
    console.print("[#B0A898]── 批次發樣／開免費單 ──[/#B0A898]")
    console.print("[dim]  一個樣品組合，一次開給多個客戶。單價全 0、狀態未出貨（之後可走出貨/出庫扣庫存）[/dim]")

    order_type = _select_with_esc("選擇單別：", choices=SAMPLE_ORDER_TYPES + ["← 返回"])
    if not order_type or order_type == "← 返回":
        return

    products = _load_sample_products()

    combo = _pick_sample_combo(products)
    if not combo:
        console.print("[#FF7700]未選組合，返回[/#FF7700]")
        return

    customers = load_customers()
    chosen = _pick_sample_customers(customers)
    if not chosen:
        console.print("[#FF7700]未選客戶，返回[/#FF7700]")
        return

    # 預覽
    console.print(f"\n[#B0A898]── 即將開立 {len(chosen)} 張「{order_type}」單（每張內容相同）──[/#B0A898]")
    console.print("  組合內容：")
    for it in combo:
        console.print(f"    {it['code']:<12} {it['name'][:20]} ×{it['qty']}  @0")
    console.print("  發給客戶：")
    for c in chosen:
        console.print(f"    {c['name']}（{c['code']}）")
    console.print(f"  單別：{order_type}　狀態：未出貨　金額：全 0")

    resolved = []
    for it in combo:
        resolved.append({"product_code": it["code"], "unit_price": 0, "quantity": it["qty"]})

    if args.dry_run:
        console.print(f"[#FF7700]★ DRY-RUN：預覽 {len(chosen)} 張單，未寫入 Ragic[/#FF7700]")
        return

    if not questionary.confirm(f"確認開立 {len(chosen)} 張「{order_type}」單？", default=True).ask():
        console.print("[#FF7700]已取消[/#FF7700]")
        return

    success = 0
    for c in chosen:
        try:
            payload = build_payload(c, resolved, order_type, "未出貨",
                                    tax_rate="", shipping_fee=0,
                                    notes="", internal_notes="批次發樣")
            result = ragic_post(SALES_ORDER_SHEET, payload)
            if result.get("status") == "SUCCESS":
                success += 1
                console.print(f"[#5A9A4A]✓ {c['name']} 開單成功[/#5A9A4A]")
                logging.info("批次發樣成功 客戶=%s 單別=%s", c["name"], order_type)
            else:
                console.print(f"[red]✗ {c['name']} 失敗：{_friendly_error(result.get('msg', str(result)))}[/red]")
                logging.warning("批次發樣失敗 客戶=%s msg=%s", c["name"], result.get("msg", result))
        except Exception as e:
            console.print(f"[red]✗ {c['name']} 發生錯誤：{_friendly_error(str(e))}[/red]")
            logging.error("批次發樣錯誤 客戶=%s error=%s", c["name"], e)
    console.print(f"[bold #5A9A4A]完成！{success}/{len(chosen)} 張「{order_type}」單已建立[/bold #5A9A4A]")
    console.print("[dim]請至 Ragic 銷貨單頁面確認[/dim]")
    _pause()


# ── 電商訂單對帳 ─────────────────────────────────────────────

def run_ecom_reconcile(args):
    """電商訂單對帳（唯讀）：讀蝦皮/官網訂單信 vs Ragic → 列漏開 + 取消待作廢。
    目前僅對帳顯示，自動開單尚未開放。"""
    try:
        from ecom import core
        from ecom.platforms import PLATFORMS
    except ImportError:
        from app.ecom import core
        from app.ecom.platforms import PLATFORMS

    BACK = "← 返回"
    label = {"shopstore": "ShopStore（官網）", "shopee": "蝦皮"}
    name = _select_with_esc("選擇電商平台：",
                            choices=[label.get(n, n) for n in PLATFORMS] + [BACK])
    if not name or name == BACK:
        return
    plat = PLATFORMS[next(n for n in PLATFORMS if label.get(n, n) == name)]

    console.print(f"\n[#B0A898]讀取 {name} 訂單信並與 Ragic 對帳中...[/#B0A898]")
    try:
        done, missing = core.reconcile(plat)
    except Exception as e:
        console.print(f"[red]✗ 對帳失敗：{e}[/red]")
        _pause()
        return

    console.print(f"\n[bold]{name} 對帳結果[/bold]　Email 訂單 {len(done) + len(missing)} ｜ "
                  f"[#5A9A4A]Ragic 已開 {len(done)}[/#5A9A4A] ｜ [#FF7700]漏開 {len(missing)}[/#FF7700]")

    for o in missing:
        flag = "  [#D14040]🔴待取貨[/#D14040]" if o.is_cod_pending else ""
        console.print(f"\n[bold]📥 {o.order_no}[/bold]  {o.date}  買家:{o.buyer or '-'}  付款:{o.pay_method}/{o.pay_status}{flag}")
        for it in o.items:
            code, prod, src = core.match_product(plat.name, it.title)
            if code:
                console.print(f"    [#5A9A4A]✓[/#5A9A4A] {code}  {(prod or {}).get('商品名稱', '')[:20]}  ×{it.qty} @ {it.price:g}")
            else:
                console.print(f"    [#D14040]✗ 對不到[/#D14040] {it.title[:24]} ×{it.qty}（需補對照表）")

    try:
        to_void, _ = core.scan_cancellations(plat)
    except Exception:
        to_void = []
    if to_void:
        console.print(f"\n[bold #D14040]🚫 {len(to_void)} 張已開單被取消 → 建議人工作廢[/bold #D14040]")
        for order_no, buyer, hit in to_void:
            console.print(f"    {order_no or '?'}  買家:{buyer or '-'}  → Ragic 備註「{hit['note']}」")

    console.print("\n[dim]（對帳顯示用，自動開單尚未開放；需開放時請告知）[/dim]")
    _pause()


# ── 主程式 ───────────────────────────────────────────────────

def main():
    _setup_logging()
    parser = argparse.ArgumentParser(description="Ragic 銷貨單自動化上傳")
    parser.add_argument("excel", nargs="?", default=None,
        help="採購單路徑（省略時自動掃描 client_order/ 下所有待處理檔案）")
    parser.add_argument("--dry-run", action="store_true", help="預覽模式，不實際送出 Ragic")
    parser.add_argument("--reset-key", action="store_true", help="重設 Ragic API Key")
    args = parser.parse_args()

    if args.reset_key:
        if _KEY_FILE.exists():
            _KEY_FILE.unlink()
            console.print("[#FF7700]已清除舊的 API Key[/#FF7700]")
        _auth_header()  # 觸發重新輸入並儲存
        return

    # 指定單一檔案模式（命令列傳入路徑）
    if args.excel:
        excel_path = Path(args.excel).expanduser().resolve()
        if not excel_path.exists():
            console.print(f"[red]找不到檔案：{excel_path}[/red]")
            sys.exit(1)
        if args.dry_run:
            console.print("[bold #FF7700]★ DRY-RUN 模式：不會實際送出，也不會移動檔案[/bold #FF7700]")
        price_index = load_price_index()
        customers   = load_customers()
        s, o, _ = process_file(excel_path, args, price_index, customers)
        console.print(f"\n[bold #5A9A4A]完成！成功處理 {s}/{o} 張訂單[/bold #5A9A4A]")
        return

    console.clear()
    _show_welcome()

    # DRY-RUN 提示（頂層，一次即可）
    if not args.dry_run:
        mode_input = questionary.text(
            "按 Enter 開始正式執行（輸入 debug 進入測試模式）：",
            default="",
        ).ask() or ""
        if mode_input.strip().lower() == "debug":
            args.dry_run = True
    if args.dry_run:
        console.print("[bold #FF7700]★ DRY-RUN 模式：不會實際送出，也不會移動檔案[/bold #FF7700]")

    # 快取懶載入（進入新建銷售單時才 API 一次）
    price_index = customers = None

    # ── 主選單 ─────────────────────────────────────────────────
    while True:
        console.print(Rule(style="#C5A059"))
        choice = _select_with_esc(
            "請選擇功能：",
            choices=[
                "新建銷售單",
                "批次發樣／開免費單（樣品/公關/贈品）",
                "建立出貨單（銷貨單拋轉）",
                "建立出庫單（出貨單拋轉）",
                "匯出庫存報表（Excel）",
                "在途查詢（採購單未到貨）",
                "電商訂單對帳（蝦皮/官網）",
                "新竹物流建單（開發中）",
                "Agent mode（AI 數據分析）",
                "退出 (Esc)",
            ],
        )

        if not choice or choice == "退出 (Esc)":
            break
        elif choice == "新建銷售單":
            if price_index is None:
                price_index = load_price_index()
                customers   = load_customers()
            run_new_sales_order(args, price_index, customers)
        elif choice == "批次發樣／開免費單（樣品/公關/贈品）":
            run_sample_orders(args)
        elif choice == "建立出貨單（銷貨單拋轉）":
            run_create_delivery_order(args)
        elif choice == "建立出庫單（出貨單拋轉）":
            run_create_outbound_order(args)
        elif choice == "匯出庫存報表（Excel）":
            BACK = "← 返回"
            sub = _select_with_esc(
                "請選擇報表類型：",
                choices=[
                    "客戶現貨報表（依模板填入單一倉現貨）",
                    "月度庫存金額統計（給會計，全倉分頁）",
                    BACK,
                ],
            )
            if not sub or sub == BACK:
                continue
            if sub.startswith("客戶現貨報表"):
                if price_index is None:
                    price_index = load_price_index()
                    customers   = load_customers()
                run_export_inventory(args, price_index)
            else:
                month = questionary.text(
                    "請輸入報表月份（YYYY-MM，留空=上個月）：",
                    default="",
                ).ask()
                try:
                    from app.export_inventory_value import export as export_inventory_value
                except ImportError:
                    from export_inventory_value import export as export_inventory_value
                try:
                    export_inventory_value((month or "").strip() or None)
                except ValueError as e:
                    console.print(f"[red]✗ {e}[/red]")
                _pause()
        elif choice == "在途查詢（採購單未到貨）":
            kw = questionary.text(
                "輸入商品編號或名稱關鍵字（留空=列出全部）：",
                default="",
            ).ask()
            try:
                from app.in_transit_query import query as query_in_transit
            except ImportError:
                from in_transit_query import query as query_in_transit
            query_in_transit((kw or "").strip() or None)
            _pause()
        elif choice == "電商訂單對帳（蝦皮/官網）":
            run_ecom_reconcile(args)
        elif choice == "新竹物流建單（開發中）":
            console.print("[#FF7700]功能開發中，敬請期待[/#FF7700]")
        elif choice == "Agent mode（AI 數據分析）":
            from ai_assistant import run_agent_mode
            run_agent_mode()

    console.print("[bold #5A9A4A]再見！[/bold #5A9A4A]")


if __name__ == "__main__":
    main()
