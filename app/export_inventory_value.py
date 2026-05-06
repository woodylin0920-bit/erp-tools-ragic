"""每月庫存金額統計：每個倉庫一個分頁，含成本與庫存現金（=數量×成本）。

用法：
    python -m app.export_inventory_value
    python -m app.export_inventory_value 2026-05    # 指定月份（影響檔名與報表標題）
"""
from __future__ import annotations

import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.ragic_upload import (
    INVENTORY_SHEET,
    PRODUCT_PRICE_SHEET,
    console,
    ragic_get,
)

BASE_OUTPUT = Path(__file__).resolve().parent.parent / "exports"
HEADERS = ["倉庫代碼", "倉庫名稱", "種類", "商品名稱", "規格", "單位", "數量", "單位成本", "庫存現金"]
EXCLUDE_WAREHOUSES = {"CH000"}  # 中國倉儲不計入庫存金額統計


def _to_int(v) -> int:
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def _parse_date(s: str) -> date:
    s = (s or "").strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return date.min


def build_cost_map(prices: dict, asof: date) -> dict:
    """以 (商品編號, 規格, 單位) → 成本 建立索引；若同 key 多筆，取最近且 ≤ asof 的生效日。"""
    buckets: dict[tuple, list] = defaultdict(list)
    for r in prices.values():
        code = str(r.get("商品編號", "")).strip()
        spec = str(r.get("規格", "")).strip()
        unit = str(r.get("單位", "")).strip()
        if not code:
            continue
        eff = _parse_date(r.get("生效日", ""))
        cost = _to_int(r.get("成本", 0))
        buckets[(code, spec, unit)].append((eff, cost))
    cost_map: dict[tuple, int] = {}
    for key, entries in buckets.items():
        valid = [e for e in entries if e[0] <= asof]
        chosen = max(valid or entries, key=lambda x: x[0])
        cost_map[key] = chosen[1]
    return cost_map


def export(month_label: str | None = None) -> Path:
    today = date.today()
    if month_label:
        title_month = month_label
        try:
            asof = datetime.strptime(month_label + "-01", "%Y-%m-%d").date()
            # 月底
            if asof.month == 12:
                asof = date(asof.year, 12, 31)
            else:
                asof = date(asof.year, asof.month + 1, 1).replace(day=1)
                from datetime import timedelta
                asof = asof - timedelta(days=1)
        except ValueError:
            asof = today
    else:
        title_month = today.strftime("%Y-%m")
        asof = today

    console.print(f"[#B0A898]載入商品單價（成本）...[/#B0A898]")
    prices = ragic_get(PRODUCT_PRICE_SHEET)
    cost_map = build_cost_map(prices, asof)
    console.print(f"[#5A9A4A]✓ 建立成本索引 {len(cost_map)} 筆[/#5A9A4A]")

    console.print(f"[#B0A898]載入倉庫庫存...[/#B0A898]")
    inventory = ragic_get(INVENTORY_SHEET)
    console.print(f"[#5A9A4A]✓ 載入 {len(inventory)} 筆庫存[/#5A9A4A]")

    by_wh: dict[str, list[dict]] = defaultdict(list)
    wh_names: dict[str, str] = {}
    missing: list[tuple] = []

    for r in inventory.values():
        wh_code = str(r.get("倉庫代碼", "")).strip()
        wh_name = str(r.get("倉庫名稱", "")).strip()
        if not wh_code or wh_code in EXCLUDE_WAREHOUSES:
            continue
        wh_names.setdefault(wh_code, wh_name)
        code = str(r.get("商品編號", "")).strip()
        spec = str(r.get("規格", "")).strip()
        unit = str(r.get("單位", "")).strip()
        qty = _to_int(r.get("數量", 0))
        cost = cost_map.get((code, spec, unit))
        if cost is None:
            missing.append((wh_code, code, spec, unit))
            cost = 0
        by_wh[wh_code].append({
            "倉庫代碼": wh_code,
            "倉庫名稱": wh_name,
            "種類": str(r.get("種類", "")).strip(),
            "商品名稱": str(r.get("商品名稱", "")).strip(),
            "規格": _to_int(spec) if spec else "",
            "單位": unit,
            "數量": qty,
            "單位成本": cost,
            "庫存現金": qty * cost,
        })

    if missing:
        console.print(f"[#FF7700]⚠ {len(missing)} 筆找不到成本，已以 0 計[/#FF7700]")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8E2D6")
    total_fill = PatternFill("solid", fgColor="F4ECD8")
    summary_rows: list[tuple[str, str, int, int]] = []

    sorted_codes = sorted(by_wh.keys(), key=lambda c: (0 if c == "TW01" else 1, c))
    for wh_code in sorted_codes:
        rows = sorted(by_wh[wh_code], key=lambda x: (x["種類"], x["商品名稱"], x["規格"] or 0))
        ws = wb.create_sheet(title=f"{wh_code}_{wh_names[wh_code]}"[:31])

        ws.cell(row=1, column=1, value=f"{title_month} 期末庫存表 — {wh_code} {wh_names[wh_code]}").font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, row in enumerate(rows, start=4):
            for c, key in enumerate(HEADERS, 1):
                cell = ws.cell(row=i, column=c, value=row[key])
                if key in ("數量", "單位成本", "庫存現金"):
                    cell.number_format = "#,##0"

        last = 3 + len(rows)
        total_qty = sum(r["數量"] for r in rows)
        total_cash = sum(r["庫存現金"] for r in rows)
        summary_rows.append((wh_code, wh_names[wh_code], total_qty, total_cash))

        total_row = last + 1
        ws.cell(row=total_row, column=6, value="合計").font = bold
        c_qty = ws.cell(row=total_row, column=7, value=total_qty)
        c_qty.font = bold
        c_qty.number_format = "#,##0"
        c_cash = ws.cell(row=total_row, column=9, value=total_cash)
        c_cash.font = bold
        c_cash.number_format = "#,##0"
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=total_row, column=c).fill = total_fill

        widths = [10, 12, 10, 36, 8, 8, 10, 10, 14]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A4"

    # 彙總分頁
    summary = wb.create_sheet(title="彙總", index=0)
    summary.cell(row=1, column=1, value=f"{title_month} 期末庫存彙總").font = Font(bold=True, size=14)
    summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    for c, h in enumerate(["倉庫代碼", "倉庫名稱", "總數量", "庫存現金"], 1):
        cell = summary.cell(row=3, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, (code, name, qty, cash) in enumerate(summary_rows, start=4):
        summary.cell(row=i, column=1, value=code)
        summary.cell(row=i, column=2, value=name)
        c_q = summary.cell(row=i, column=3, value=qty)
        c_q.number_format = "#,##0"
        c_c = summary.cell(row=i, column=4, value=cash)
        c_c.number_format = "#,##0"
    total_row = 3 + len(summary_rows) + 1
    summary.cell(row=total_row, column=2, value="合計").font = bold
    tq = summary.cell(row=total_row, column=3, value=sum(s[2] for s in summary_rows))
    tq.font = bold
    tq.number_format = "#,##0"
    tc = summary.cell(row=total_row, column=4, value=sum(s[3] for s in summary_rows))
    tc.font = bold
    tc.number_format = "#,##0"
    for c in range(1, 5):
        summary.cell(row=total_row, column=c).fill = total_fill
    for c, w in enumerate([10, 16, 12, 16], 1):
        summary.column_dimensions[get_column_letter(c)].width = w

    BASE_OUTPUT.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = BASE_OUTPUT / f"inventory_value_{title_month.replace('-', '')}_{stamp}.xlsx"
    wb.save(out_path)
    console.print(f"[#5A9A4A]✓ 已輸出：{out_path}[/#5A9A4A]")
    grand_total = sum(s[3] for s in summary_rows)
    console.print(f"[#5A9A4A]  全公司庫存現金合計：{grand_total:,}[/#5A9A4A]")
    return out_path


if __name__ == "__main__":
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    export(arg)
