"""
TRUParser - 玩具反斗城報價表格式
格式特徵：
  - 一張 Excel = 多張 Ragic 訂單（每個門市各一張）
  - Row 2-3：表頭，門市代碼（4402, 4418 ...）在 Column P (index 15) 之後
  - 商品列（Column D = 13碼條碼）從 Row 4 起
  - Column J = 單價（不含稅）
  - PO 號碼：Column N (index 13)
  - 每個門市欄值 > 0 的商品組成該門市的訂單
  - 多 PO 支援：若有 PO# 欄位，段落標題列會在該欄標記新 PO 號，
    後續商品列沿用此 PO，每個 (門市, PO) 組合建立一張獨立訂單
"""

import io
import re
import zipfile
from typing import Dict, List, Optional, Tuple

import openpyxl

from .base import BaseParser, Order, OrderItem

def _strip_autofilter(xml: str) -> str:
    # Remove block form: <autoFilter ...> ... </autoFilter>
    xml = re.sub(r'<autoFilter\b.*?</autoFilter>', '', xml, flags=re.DOTALL | re.IGNORECASE)
    # Remove self-closing form: <autoFilter ... />
    xml = re.sub(r'<autoFilter\b[^>]*/>', '', xml, flags=re.IGNORECASE)
    return xml


def _load_wb(filepath: str):
    """Load workbook, stripping broken <autoFilter> XML on first failure."""
    try:
        return openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        buf = io.BytesIO()
        with zipfile.ZipFile(filepath, 'r') as zin, \
             zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                    data = _strip_autofilter(data.decode('utf-8')).encode('utf-8')
                zout.writestr(item, data)
        buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)

BARCODE_COL    = 3   # Column D (0-indexed)
UNIT_PRICE_COL = 9   # Column J
PO_COL         = 13  # Column N
STORE_START_COL = 15 # Column P (first store column)

BARCODE_RE = re.compile(r'^\d{12,14}$')
PO_RE      = re.compile(r'^\d{5,8}$')   # PO 號格式：5-8 位數字

# 商品/固定欄位的標題關鍵字（門市欄之前的欄位）。用來動態定位「門市區起點」=
# 第一個店號左側、最右邊的固定欄位 +1，避免寫死欄位位置（TRU 改版/跑版時自動跟著走）。
# 注意：不含「各店/PO#/門市」標記與門市名稱本身。
FIXED_HEADER_KEYS = (
    "IP", "系列", "英文", "條碼", "參考", "規格", "零售價", "MSRP",
    "定價", "折數", "單價", "備註", "數量", "SKN", "PO號碼",
    "現貨", "訂單總數", "台灣現貨", "廠商自填",
)


def _val(cell) -> str:
    return str(cell).strip() if cell is not None else ""


def _float(cell) -> float:
    try:
        return float(cell) if cell is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


class TRUParser(BaseParser):

    def parse(self) -> List[Order]:
        wb = _load_wb(self.filepath)
        orders = []

        for sheet in wb.worksheets:
            orders.extend(self._parse_sheet(sheet))

        return orders

    def _parse_sheet(self, sheet) -> List[Order]:
        all_rows = list(sheet.iter_rows(values_only=True))
        if len(all_rows) < 3:
            return []

        # Find the header row that contains store codes (4-digit numbers like 4402)
        header_row_idx, store_cols, po_override_col = self._find_store_columns(all_rows)
        if not store_cols:
            return []

        # Build orders keyed by (store_code, po_number) to support multi-PO files
        store_orders: Dict[Tuple[str, str], Order] = {}

        # 依「欄位標題」自動定位關鍵欄，TRU 改版/跑版時不必再手動改欄位位置。
        # 找不到標題才退回資料驅動偵測或預設常數。
        barcode_col = self._resolve_col(all_rows, header_row_idx, ["條碼"],
                                        self._detect_barcode_col(all_rows[header_row_idx + 1:]))
        price_col   = self._resolve_col(all_rows, header_row_idx, ["單價"], UNIT_PRICE_COL)
        name_col    = self._resolve_col(all_rows, header_row_idx, ["系列", "品名", "品名稱"], 1)
        # 真正的「PO號碼」欄（依標題定位），避免誤抓到「現貨」欄；再算整檔主要 PO（眾數）後援，
        # 確保個別列 PO 留白時「一店一單」不被拆散。
        po_data_col = self._resolve_col(all_rows, header_row_idx, ["PO號碼"], PO_COL)
        master_po   = self._modal_po(all_rows[header_row_idx + 1:], po_data_col)

        for row in all_rows[header_row_idx + 1:]:
            if len(row) <= barcode_col:
                continue

            barcode = _val(row[barcode_col])
            if not BARCODE_RE.match(barcode):
                continue

            # Determine PO per column:
            #   - stores LEFT  of po_override_col → PO from po_data_col
            #   - stores RIGHT of po_override_col → PO from po_override_col cell
            po_left  = _val(row[po_data_col]) if po_data_col < len(row) else ""
            po_right = (
                _val(row[po_override_col])
                if po_override_col is not None and po_override_col < len(row)
                else ""
            )
            unit_price = _float(row[price_col]) if price_col < len(row) else 0.0
            le_name    = _val(row[name_col]) if name_col < len(row) else ""

            for store_code, col_idx in store_cols.items():
                if col_idx >= len(row):
                    continue
                qty = _float(row[col_idx])
                if qty <= 0:
                    continue

                # Pick PO based on which side of po_override_col this store is。
                # 後援順序：逐列 PO → 整檔主要 PO → sheet 名稱（確保同店所有品項歸同一張單）。
                if po_override_col is not None and col_idx > po_override_col:
                    po_number = po_right or po_left or master_po or sheet.title
                else:
                    po_number = po_left or master_po or sheet.title

                key = (store_code, po_number)
                if key not in store_orders:
                    store_orders[key] = Order(
                        client_code="TRU",
                        store_code=store_code,
                        po_number=po_number,
                        source_file=self.filepath,
                    )

                store_orders[key].items.append(OrderItem(
                    barcode=barcode,
                    quantity=qty,
                    le_name=le_name,
                    unit_price=unit_price,
                ))

        # Sort by Excel column order: left-of-PO# stores first, then right-of-PO# stores
        col_order = {code: i for i, code in enumerate(store_cols.keys())}

        def _sort_key(order):
            col = col_order.get(order.store_code, 999)
            group = 1 if (po_override_col is not None and col > po_override_col) else 0
            return (group, col)

        return sorted(store_orders.values(), key=_sort_key)

    def _resolve_col(self, all_rows, header_row_idx, keywords, default) -> int:
        """依標題關鍵字自動定位欄位。掃描標題區（header 列及其上方數列），
        回傳第一個標題含任一關鍵字的欄 index；找不到時回 default。
        如此 TRU 改版/跑版（欄位左右移動）時不需再手動改欄位常數。"""
        for r in range(0, min(header_row_idx + 1, len(all_rows))):
            for col_idx, cell in enumerate(all_rows[r]):
                t = _val(cell)
                if t and any(k in t for k in keywords):
                    return col_idx
        return default

    def _detect_barcode_col(self, data_rows) -> int:
        """資料驅動偵測條碼欄：取資料列中、最多列符合 12-14 碼數字的欄位。
        作為標題定位失敗時的後援。找不到時回預設 BARCODE_COL。"""
        counts: Dict[int, int] = {}
        for row in data_rows:
            for c, cell in enumerate(row):
                if BARCODE_RE.match(_val(cell)):
                    counts[c] = counts.get(c, 0) + 1
        return max(counts, key=counts.get) if counts else BARCODE_COL

    def _modal_po(self, data_rows, po_data_col) -> str:
        """取該欄出現次數最多、且符合 PO 格式（5-8 碼數字）的值，作為整檔主要 PO。"""
        counts: Dict[str, int] = {}
        for row in data_rows:
            if po_data_col >= len(row):
                continue
            v = _val(row[po_data_col])
            if PO_RE.match(v):
                counts[v] = counts.get(v, 0) + 1
        return max(counts, key=counts.get) if counts else ""

    def _find_store_columns(self, all_rows) -> tuple:
        """
        Scan rows 0-4 for the row containing 4-digit TRU store codes (e.g. 4402).
        Also capture text-named stores from the same row AND the row above
        (e.g. DC in row above, 統領/板橋大遠百 in same row).
        Returns (header_row_index, {store_code: col_index}, po_override_col).
        po_override_col: column index of the 'PO#' special column, or None.
        """
        store_code_re = re.compile(r'^4\d{3}$')
        skip_vals = {"TTL", "TOTAL", "合計", "小計", "PO號碼", "PO", "PO#", "",
                     # 報價/補貨表表頭欄位，非門市，需排除（否則會誤建幽靈訂單）
                     "訂單總數", "現貨", "各店", "訂單", "TTL QTY", "台灣現貨"}

        for row_idx, row in enumerate(all_rows[:5]):
            # 找出本列所有 4 碼門市店號欄（如 4402）。需 ≥2 個才視為門市標題列，
            # 避免偶發單一 4 碼數字誤判。
            code_cols = [c for c, cell in enumerate(row) if store_code_re.match(_val(cell))]
            if len(code_cols) < 2:
                continue

            prev_row = all_rows[row_idx - 1] if row_idx > 0 else []

            # 門市區起點：以第一個店號欄為錨，往左最多 3 欄找「各店/PO#/門市」標記；
            # 找不到就從第一個店號欄起。如此 TRU 整體跑版時門市區仍能正確定位，
            # 不會把左側的 SKN／現貨／PO號碼 等欄誤當門市（舊版寫死 STORE_START_COL 會出錯）。
            first_code = min(code_cols)
            # 門市區起點 = 第一個店號左側、最右邊的「固定欄位標題」+1。
            # 如此位於店號左邊的文字門市（如 DC）仍會納入，且固定欄位多寡/位置改變時自動跟著走。
            fixed_max = -1
            for r in range(0, row_idx + 1):
                hr = all_rows[r]
                for c in range(min(first_code, len(hr))):
                    t = _val(hr[c])
                    if t and any(k in t for k in FIXED_HEADER_KEYS) and c > fixed_max:
                        fixed_max = c
            floor = fixed_max + 1 if fixed_max >= 0 else (STORE_START_COL - 1)

            # Collect all stores in LEFT-TO-RIGHT column order。
            # Check both this row and the row above for each column
            # (e.g. DC 在上一列、統領/板橋大遠百 在同列)。
            store_cols = {}
            po_override_col = None
            for col_idx in range(floor, len(row)):
                val = _val(row[col_idx]) if col_idx < len(row) else ""
                if not val and col_idx < len(prev_row):
                    val = _val(prev_row[col_idx])

                if not val:
                    continue
                if val.upper() in ("PO#", "各店", "各店PO#"):
                    po_override_col = col_idx   # PO# 標記欄，不加入門市
                    continue
                if val.upper() not in skip_vals:
                    store_cols[val] = col_idx

            return row_idx, store_cols, po_override_col

        return 0, {}, None
