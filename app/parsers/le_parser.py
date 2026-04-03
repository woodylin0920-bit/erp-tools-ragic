"""
LEParser - 麗嬰國際採購單格式
格式特徵：
  - 每個工作表 = 一張門市訂單
  - 前 ~50 行是表頭（含門市代碼 AD227、PO 號 PO-1256789）
  - 從 Row 53 起，Column L (index 11) = 13碼條碼，Column S (index 18) = 應到數
"""

import re
from typing import List

import openpyxl

from .base import BaseParser, Order, OrderItem

BARCODE_COL = 11   # Column L (0-indexed)
QTY_COL     = 18   # Column S (0-indexed)
LE_CODE_COL = 1    # Column B
LE_NAME_COL = 5    # Column F

BARCODE_RE    = re.compile(r'^\d{12,14}$')
STORE_CODE_RE = re.compile(r'\b([A-Z]{2}\d{3})\b')
PO_RE         = re.compile(r'(PO[-‐]\d+)', re.IGNORECASE)


def _scan_cell(cell_value) -> str:
    return str(cell_value).strip() if cell_value is not None else ""


class LEParser(BaseParser):

    def parse(self) -> List[Order]:
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        orders = []

        for sheet in wb.worksheets:
            store_code, po_number = self._extract_metadata(sheet)
            items = self._extract_items(sheet)

            if items:
                orders.append(Order(
                    client_code="LE",
                    store_code=store_code or "???",
                    po_number=po_number or sheet.title,
                    items=items,
                    source_file=self.filepath,
                ))

        return orders

    def _extract_metadata(self, sheet) -> tuple:
        """Scan first 55 rows for store code and PO number."""
        store_code = None
        po_number = None

        for row in sheet.iter_rows(min_row=1, max_row=55, values_only=True):
            for cell in row:
                text = _scan_cell(cell)
                if not text:
                    continue
                if not store_code:
                    m = STORE_CODE_RE.search(text)
                    if m:
                        store_code = m.group(1)
                if not po_number:
                    m = PO_RE.search(text)
                    if m:
                        po_number = m.group(1)
                if store_code and po_number:
                    return store_code, po_number

        return store_code, po_number

    def _extract_items(self, sheet) -> List[OrderItem]:
        items = []
        max_col = sheet.max_column or 25

        for row in sheet.iter_rows(min_row=1, values_only=True):
            if max_col < BARCODE_COL + 1:
                continue
            if len(row) <= max(BARCODE_COL, QTY_COL):
                continue

            barcode = _scan_cell(row[BARCODE_COL])
            if not BARCODE_RE.match(barcode):
                continue

            qty_raw = row[QTY_COL] if QTY_COL < len(row) else None
            try:
                qty = float(qty_raw) if qty_raw is not None else 0.0
            except (ValueError, TypeError):
                qty = 0.0

            if qty <= 0:
                continue

            le_code = _scan_cell(row[LE_CODE_COL]) if LE_CODE_COL < len(row) else ""
            le_name = _scan_cell(row[LE_NAME_COL]) if LE_NAME_COL < len(row) else ""

            items.append(OrderItem(
                barcode=barcode,
                quantity=qty,
                le_code=le_code,
                le_name=le_name,
            ))

        return items
