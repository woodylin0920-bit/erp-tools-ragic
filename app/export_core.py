"""匯出庫存報表的純邏輯（無互動）。讀 Ragic 庫存、填客戶模板、輸出 Excel。
★ 不寫入 Ragic，只產生本機檔案。★ 供 GUI 使用；CLI 仍走 ragic_upload 既有實作。
"""
import re
import zipfile
from datetime import datetime

import ragic_upload as R

BULK_UNITS = {"中盒", "箱", "整箱", "端盒"}


def load_warehouses() -> dict:
    """{倉庫代碼: 倉庫名稱}。"""
    inv = R.ragic_get(R.INVENTORY_SHEET)
    whs = {}
    for rec in inv.values():
        wh = str(rec.get("倉庫代碼", "")).strip()
        if wh:
            whs[wh] = str(rec.get("倉庫名稱", "")).strip()
    return whs


def list_templates() -> list:
    """templates/ 下的 .xlsx 模板路徑（新到舊）。"""
    R.BASE_TEMPLATES.mkdir(exist_ok=True)
    return sorted(R.BASE_TEMPLATES.glob("*.xlsx"), reverse=True)


def export_to_template(warehouse_code: str, template_path, price_index: dict):
    """讀庫存→換算 PCS（只算中盒/箱類）→填模板「現貨」欄→輸出 exports/。
    回 (out_path, filled, skipped)。找不到「現貨」欄會 raise ValueError。"""
    import openpyxl
    from openpyxl.workbook.properties import CalcProperties

    R.BASE_OUTPUT.mkdir(exist_ok=True)
    inventory_all = R.ragic_get(R.INVENTORY_SHEET)

    code_to_barcode = {}
    for barcode, entries in price_index.items():
        for e in entries:
            code_to_barcode[re.sub(r"-\d+$", "", e["product_code"])] = barcode

    inventory_pcs, skipped = {}, 0
    for rec in inventory_all.values():
        if str(rec.get("倉庫代碼", "")).strip() != warehouse_code:
            continue
        if str(rec.get("單位", "")).strip() not in BULK_UNITS:
            skipped += 1
            continue
        prod = str(rec.get("商品編號", "")).strip()
        try:
            qty = int(float(rec.get("數量", 0) or 0))
        except (ValueError, TypeError):
            qty = 0
        try:
            spec = int(float(rec.get("規格", 1) or 1))
        except (ValueError, TypeError):
            spec = 1
        bc = code_to_barcode.get(prod)
        if bc:
            inventory_pcs[bc] = inventory_pcs.get(bc, 0) + qty * spec

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    inv_col = None
    for r in (2, 3):
        for cell in ws[r]:
            if str(cell.value or "").strip() == "現貨":
                inv_col = cell.column - 1
                break
        if inv_col is not None:
            break
    if inv_col is None:
        raise ValueError("此模板找不到「現貨」欄位，請選 inventory 或 quote 模板")

    filled = 0
    for row in ws.iter_rows(min_row=4):
        d = row[3]
        if d.value is None:
            continue
        try:
            bc = str(int(float(d.value)))
        except (ValueError, TypeError):
            continue
        if bc in inventory_pcs and inv_col < len(row):
            row[inv_col].value = inventory_pcs[bc]
            filled += 1

    wb.calculation = CalcProperties(fullCalcOnLoad=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    prefix = template_path.stem.replace("-template", "")
    out_path = R.BASE_OUTPUT / f"{prefix}_{warehouse_code}_{ts}.xlsx"
    wb.save(out_path)

    # 保留模板嵌入圖片（openpyxl 3.x 存檔會掉圖）：注入模板的 media 等，只換邏輯內容
    with zipfile.ZipFile(template_path) as zt:
        merged = {n: zt.read(n) for n in zt.namelist()}
    with zipfile.ZipFile(out_path) as zo:
        oxl = {n: zo.read(n) for n in zo.namelist()}
    for f in ("xl/worksheets/sheet1.xml", "xl/sharedStrings.xml",
              "xl/styles.xml", "xl/workbook.xml"):
        if f in oxl:
            merged[f] = oxl[f]
    merged.pop("xl/calcChain.xml", None)
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in merged.items():
            zf.writestr(name, data)

    return out_path, filled, skipped
